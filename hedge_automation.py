import streamlit as st
import pandas as pd
import os
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Pillow / image support (graceful if missing)
try:
    from openpyxl.drawing.image import Image as XLImage
    PIL_OK = True
except ImportError:
    XLImage = None
    PIL_OK = False

# Initialize session state at module level
if 'user_ids' not in st.session_state:
    st.session_state.user_ids = []
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
if 'stats' not in st.session_state:
    st.session_state.stats = None
if 'download_name' not in st.session_state:
    st.session_state.download_name = None

def clean_header(headers):
    """Cleans and normalizes DataFrame headers."""
    return [str(h).strip().replace("\n", " ") for h in headers]

def process_csv(uploaded_file, filename: str) -> pd.DataFrame:
    """Read CSV from Streamlit's UploadedFile object, normalize headers, and add Server column."""
    if not filename:
        raise ValueError("Cannot determine server name from an empty filename.")

    # Read headers first to validate
    headers_df = pd.read_csv(uploaded_file, nrows=0)
    uploaded_file.seek(0)  # Reset file pointer after reading headers
    headers = clean_header(list(headers_df.columns))

    if len(headers) not in [19, 20]:
        raise ValueError(f"Error: Expected 19 or 20 headers, but found {len(headers)}.")

    if len(headers) == 20:
        headers = headers[:19] + ["Tag"]
    else:  # len is 19
        headers = headers + ["Tag"]

    # Read the full data
    df = pd.read_csv(uploaded_file, header=None, skiprows=1, dtype=str)
    if len(df.columns) != 20:
        raise ValueError(f"Error: Expected 20 columns in data rows, but found {len(df.columns)}.")

    df.columns = headers

    # Add Server column from the original filename
    server_name = filename.split('_')[0].split(' ')[0]
    df['Server'] = server_name

    return df

def _maybe_to_number(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    s = str(val).strip()
    if s == "": return ""
    s_clean = s.replace(",", "")
    if not re.compile(r'^[+-]?\d+([.]\d+)?$').match(s_clean): return val
    if "." not in s_clean and "e" not in s_clean.lower():
        if len(s_clean.lstrip("+-")) > 15: return s
        try: return int(s_clean)
        except (ValueError, TypeError): return val
    try: return float(s_clean)
    except (ValueError, TypeError): return val

def _set_num_format(cell, value):
    if isinstance(value, int): cell.number_format = "0"
    elif isinstance(value, float): cell.number_format = "0.00"

def _autofit(ws, scan_rows=200, min_w=10, max_w=60, skip_letters=None):
    skip = set(skip_letters or [])
    for col in ws.columns:
        try:
            letter = get_column_letter(col[0].column)
            if letter in skip: continue
            max_len = 0
            for cell in col[:scan_rows]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(min_w, max_len + 2), max_w)
        except Exception: pass

def export_orderbook_with_formulas(df: pd.DataFrame, user_ids: list[str], mtm_image_file: io.BytesIO | None = None) -> tuple[Workbook, dict]:
    """Create the HEDGES workbook object; returns (Workbook, stats_dict)."""
    if 'Status' in df.columns:
        df['Status'] = df['Status'].fillna('').str.upper()
        status_col = 'Status'
    elif 'Tag' in df.columns:
        df['Tag'] = df['Tag'].fillna('').str.upper()
        status_col = 'Tag'
    else:
        raise ValueError("No 'Status' or 'Tag' column found in CSV.")

    formula_columns = {
        "C": ["=RIGHT(B2,2)", "=RIGHT(B{r},2)"],
        "K": ["=IF(F2=\"SELL\",-H2,H2)", "=IF(F{r}=\"SELL\",-H{r},H{r})"],
        "M": ["=IF(C2=\"CE\",K2*1,K2*0)", "=IF(C{r}=\"CE\",K{r}*1,K{r}*0)"],
        "N": ["=IF(C2=\"PE\",K2*1,K2*0)", "=IF(C{r}=\"PE\",K{r}*1,K{r}*0)"],
        "O": ["=M2", "=O{prev}+M{r}"],
        "P": ["=N2", "=P{prev}+N{r}"],
        "S": ["=IF(M2<0,0,M2)", "=IF(M{r}<0,0,M{r})+S{prev}"],
        "T": ["=IF(M2<0,M2,0)", "=IF(M{r}<0,M{r},0)+T{prev}"],
        "U": ["=IF(P2<0,0,P2)", "=IF(N{r}<0,0,N{r})+U{prev}"],
        "V": ["=IF(N2<0,N2,0)", "=IF(N{r}<0,N{r},0)+V{prev}"],
        "W": ["=IFERROR(ABS(S2)/ABS(T2),0)", "=IFERROR(ABS(S{r})/ABS(T{r}),0)"],
        "X": ["=IFERROR(ABS(U2)/ABS(V2),0)", "=IFERROR(ABS(U{r})/ABS(V{r}),0)"]
    }
    formula_cols_letters = list(formula_columns.keys())

    custom_headers = [
        "SNO", "Symbol", "CE/PE", "Order Time", "Order ID", "Transaction", "Order Type",
        "Quantity", "Price", "Exchange Time", "B/S", "Avg Price", "CE", "PE",
        "CECum", "PECum", "User Alias", "User ID",
        "CE Buy", "CE Sell", "PE Buy", "PE Sell", "CE B/S", "PE B/S"
    ]

    wb = Workbook()
    wb.remove(wb.active)

    ws_full = wb.create_sheet("Orderbook Full")
    for c_idx, header in enumerate(df.columns, start=1):
        ws_full.cell(row=1, column=c_idx).value = header
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            val = _maybe_to_number(value)
            cell = ws_full.cell(row=r_idx, column=c_idx)
            cell.value = val
            _set_num_format(cell, val)
    _autofit(ws_full)

    created_sheets = []
    for uid in [u.strip() for u in user_ids if u.strip()]:
        user_df = df[(df.get('User ID', pd.Series(dtype=str)) == uid) & (df[status_col] == 'COMPLETE')]
        ws = wb.create_sheet(uid)
        created_sheets.append(uid)

        for c_idx, header in enumerate(custom_headers, start=1):
            ws.cell(row=1, column=c_idx).value = header

        start_row = 2
        for r_off, row in enumerate(user_df.itertuples(index=False), start=0):
            r = start_row + r_off
            for c_idx, value in enumerate(row, start=1):
                val = _maybe_to_number(value)
                cell = ws.cell(row=r, column=c_idx)
                cell.value = val
                _set_num_format(cell, val)
            for col_letter, (f_first, f_next) in formula_columns.items():
                formula = f_first if r == 2 else f_next.replace('{r}', str(r)).replace('{prev}', str(r - 1))
                ws[f"{col_letter}{r}"] = formula

    ws_mtm = wb.create_sheet("MTM", index=len(wb.worksheets))
    ws_mtm["A1"] = "MTM Image:"
    ws_mtm["A1"].font = ws_mtm["A1"].font.copy(bold=True)

    if mtm_image_file and PIL_OK:
        try:
            img = XLImage(mtm_image_file)
            target_width = 800
            if hasattr(img, 'width') and img.width:
                scale = target_width / float(img.width)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            ws_mtm.add_image(img, "A2")
        except Exception as e:
            ws_mtm["A3"] = f"(Could not load image: {e})"
    elif mtm_image_file and not PIL_OK:
        ws_mtm["A3"] = "(Pillow library not installed; cannot embed image.)"
    else:
        ws_mtm["A3"] = "(No image uploaded)"

    light_green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    FORMULA_WIDTHS = {"C": 6, "K": 7, "M": 7, "N": 7, "O": 8, "P": 8, "S": 8, "T": 8, "U": 8, "V": 8, "W": 10, "X": 10}
    
    for sheet_name in created_sheets:
        ws = wb[sheet_name]
        last_row = ws.max_row
        for col_letter in formula_cols_letters:
            for r in range(2, last_row + 1):
                ws[f"{col_letter}{r}"].fill = light_green
        for col_letter, width in FORMULA_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        _autofit(ws, scan_rows=300, min_w=12, max_w=60, skip_letters=formula_cols_letters)

    stats = {
        "total_rows": len(df),
        "total_cols": len(df.columns),
        "sheets": created_sheets,
    }

    return wb, stats

def run():
    # if st.button("🔙 Back to Dashboard", key="back_dashboard"):
    #     st.session_state.current_page = 'dashboard'
    #     st.rerun()
    # Fallback session state initialization
    if 'user_ids' not in st.session_state:
        st.session_state.user_ids = []
    if 'excel_data' not in st.session_state:
        st.session_state.excel_data = None
    if 'stats' not in st.session_state:
        st.session_state.stats = None
    if 'download_name' not in st.session_state:
        st.session_state.download_name = None

    # Custom CSS for improved styling
    st.markdown("""
        <style>
        :root {
            --bg-light: #F9FAFB;
            --text-light: #1F2937;
            --accent: #3B82F6;
            --accent-hover: #2563EB;
            --border-light: #E5E7EB;
            --shadow-light: rgba(0, 0, 0, 0.1);
        }
        .main-container {
            font-family: 'Inter', sans-serif;
            max-width: 800px;
            margin: 2rem auto;
            padding: 2rem;
            border-radius: 0.75rem;
            box-shadow: 0 4px 8px var(--shadow-light);
            background: var(--bg-light);
            color: var(--text-light);
        }
        .main-container .header {
            font-size: 2.25rem;
            font-weight: 800;
            background: linear-gradient(to right, var(--accent), #6366F1);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-align: center;
        }
        .main-container .subheader {
            font-size: 1.125rem;
            color: #4B5563;
            text-align: center;
            margin-bottom: 1.5rem;
        }
        .main-container .stButton > button {
            background: var(--accent);
            color: white;
            border: none;
            padding: 0.75rem 1.5rem;
            border-radius: 0.5rem;
            font-weight: 600;
            transition: all 0.3s ease;
            width: 100%;
        }
        .main-container .stButton > button:hover {
            background: var(--accent-hover);
            transform: translateY(-2px);
            box-shadow: 0 4px 6px var(--shadow-light);
        }
        .main-container .stTextInput > div > div > input {
            border: 1px solid var(--border-light);
            border-radius: 0.5rem;
            padding: 0.75rem;
        }
        .main-container .stFileUploader > div > div > input {
            border: 1px solid var(--border-light);
            border-radius: 0.5rem;
            padding: 0.75rem;
        }
        .main-container .user-id-container {
            border: 1px solid var(--border-light);
            border-radius: 0.5rem;
            padding: 1rem;
            margin-bottom: 1.5rem;
        }
        .main-container .fade-in {
            animation: fadeIn 0.5s ease-out;
        }
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @media (max-width: 640px) {
            .main-container .header {
                font-size: 1.875rem;
            }
            .main-container .st-columns > div {
                margin-bottom: 1rem;
            }
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<h1 class="header fade-in">📈 Hedges Generator</h1>', unsafe_allow_html=True)
    st.markdown('<p class="subheader fade-in">Create Excel exports from your Orderbook CSV for automated hedge operations.</p>', unsafe_allow_html=True)

    # User ID Management
    st.markdown('<h2 class="text-lg font-semibold mt-6 mb-3 fade-in">1. Enter User IDs</h2>', unsafe_allow_html=True)
    st.markdown('<p class="subheader fade-in">Add User IDs to create individual sheets in the Excel file.</p>', unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        new_user_id = st.text_input("new_user_id", placeholder="Enter a User ID", label_visibility="collapsed")
    with col2:
        if st.button("Add User ID", use_container_width=True, key="add_user_id"):
            if new_user_id:
                user_to_add = new_user_id.strip().upper()
                if user_to_add not in st.session_state.user_ids:
                    st.session_state.user_ids.append(user_to_add)
                    st.toast(f"✅ Added '{user_to_add}'!", icon="🎉")
                    st.rerun()
                else:
                    st.toast(f"⚠️ '{user_to_add}' is already in the list.", icon="🚨")
            else:
                st.toast("Please enter a User ID.", icon="❗")

    # Fallback check before accessing user_ids
    if 'user_ids' not in st.session_state:
        st.session_state.user_ids = []
    if st.session_state.user_ids:
        st.markdown('<p class="subheader fade-in">Current User IDs:</p>', unsafe_allow_html=True)
        with st.container(border=True, height=150):
            for i, user_id in enumerate(st.session_state.user_ids):
                col1, col2 = st.columns([4, 1])
                col1.code(user_id, language=None)
                if col2.button("❌", key=f"remove_{user_id}", help=f"Remove {user_id}", use_container_width=True):
                    st.session_state.user_ids.pop(i)
                    st.toast(f"🗑️ Removed '{user_id}'.")
                    st.rerun()
    else:
        st.info("No User IDs have been added yet.", icon="ℹ️")

    # Form for File Uploads and Submission
    with st.form("hedges_form"):
        st.markdown('<h2 class="text-lg font-semibold mt-6 mb-3 fade-in">2. Upload Orderbook CSV</h2>', unsafe_allow_html=True)
        uploaded_csv = st.file_uploader(
            "Select a CSV file",
            type="csv",
            help="Upload the orderbook CSV file. Ensure it has 19 or 20 columns including 'Status' or 'Tag' and 'User ID'.",
            key="csv_uploader"
        )

        st.markdown('<h2 class="text-lg font-semibold mt-6 mb-3 fade-in">3. Upload MTM Image (Optional)</h2>', unsafe_allow_html=True)
        if not PIL_OK:
            st.warning("Pillow library is not installed. Image uploads are disabled. Run `pip install pillow` to enable.", icon="⚠️")
            uploaded_image = None
        else:
            uploaded_image = st.file_uploader(
                "Select an image file for the MTM sheet",
                type=["png", "jpg", "jpeg", "bmp", "gif"],
                help="Upload an optional MTM image to include in the Excel file.",
                key="image_uploader"
            )

        st.markdown('<hr class="my-6">', unsafe_allow_html=True)
        submitted = st.form_submit_button("🚀 Generate HEDGES Excel", use_container_width=True)

    # Processing Logic
    if submitted:
        st.session_state.excel_data = None
        st.session_state.stats = None
        st.session_state.download_name = None

        if not uploaded_csv:
            st.error("Please upload an Orderbook CSV file.", icon="❌")
        elif not st.session_state.user_ids:
            st.error("Please add at least one User ID.", icon="❌")
        else:
            user_ids = st.session_state.user_ids
            with st.spinner("Processing... Generating your Excel file..."):
                try:
                    df = process_csv(uploaded_csv, uploaded_csv.name)
                    workbook, stats = export_orderbook_with_formulas(df, user_ids, uploaded_image)
                    excel_buffer = io.BytesIO()
                    workbook.save(excel_buffer)
                    st.session_state.excel_data = excel_buffer.getvalue()
                    st.session_state.stats = stats
                    base_name = os.path.splitext(uploaded_csv.name)[0]
                    st.session_state.download_name = f"{base_name} HEDGES.xlsx"
                except Exception as e:
                    st.error(f"An error occurred: {e}", icon="❌")
                    st.exception(e)

    # Display Results
    if st.session_state.excel_data:
        st.success("✅ Excel file generated successfully!", icon="🎉")
        st.markdown(f'<p class="subheader fade-in">Your file <strong>{st.session_state.download_name}</strong> is ready to download.</p>', unsafe_allow_html=True)

        col1, col2, col3 = st.columns(3)
        col1.metric("Total Rows Processed", f"{st.session_state.stats['total_rows']:,}")
        col2.metric("Columns Detected", st.session_state.stats['total_cols'])
        col3.metric("User Sheets Created", len(st.session_state.stats['sheets']))

        st.download_button(
            label="⬇️ Download Excel File",
            data=st.session_state.excel_data,
            file_name=st.session_state.download_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel",
            use_container_width=True
        )

    st.markdown('</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    run()
