# # # import streamlit as st
# # # import pandas as pd
# # # import numpy as np
# # # import re
# # # import base64
# # # from io import BytesIO
# # # import logging
# # # import openpyxl
# # # from openpyxl.styles import Alignment, Border, Side, Font
# # # from datetime import datetime

# # # # Set up logging
# # # logging.basicConfig(level=logging.INFO)
# # # logger = logging.getLogger(__name__)

# # # # Check Streamlit version for compatibility
# # # try:
# # #     import streamlit
# # #     logger.info(f"Streamlit version: {streamlit.__version__}")
# # # except ImportError:
# # #     st.error("Streamlit is not installed. Please install it using `pip install streamlit`.")
# # #     st.stop()

# # # def run():
# # #     # Main title and description
# # #     st.markdown("# Realized PNL for 19")
# # #     st.write("This tool displays realized profit and loss for 19.")

# # #     # Enhanced Custom CSS with Tailwind CSS and Animations
# # #     st.markdown("""
# # #         <link href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css" rel="stylesheet">
# # #         <style>
# # #         body {
# # #             font-family: 'Inter', sans-serif;
# # #             background: linear-gradient(135deg, #f3f4f6, #e5e7eb);
# # #         }
# # #         .stButton>button {
# # #             background: linear-gradient(45deg, #3b82f6, #60a5fa);
# # #             color: white;
# # #             border: none;
# # #             padding: 0.75rem 1.5rem;
# # #             border-radius: 0.5rem;
# # #             font-weight: 600;
# # #             transition: all 0.3s ease;
# # #             width: 100%;
# # #         }
# # #         .stButton>button:hover {
# # #             background: linear-gradient(45deg, #2563eb, #3b82f6);
# # #             transform: translateY(-2px);
# # #             box-shadow: 0 4px 6px rgba(0,0,0,0.2);
# # #         }
# # #         .stDateInput input {
# # #             border: 2px solid #3b82f6;
# # #             border-radius: 0.5rem;
# # #             padding: 0.5rem;
# # #             background: #ffffff;
# # #             color: #1f2937;
# # #             font-size: 1rem;
# # #             transition: all 0.3s ease;
# # #         }
# # #         .stDateInput input:focus {
# # #             outline: none;
# # #             border-color: #2563eb;
# # #             box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
# # #             background: #f8fafc;
# # #         }
# # #         .stFileUploader button {
# # #             background: linear-gradient(45deg, #10b981, #34d399);
# # #             color: white;
# # #             border-radius: 0.5rem;
# # #             padding: 0.75rem;
# # #         }
# # #         .stFileUploader button:hover {
# # #             background: linear-gradient(45deg, #059669, #10b981);
# # #         }
# # #         .stCheckbox label {
# # #             font-size: 1rem;
# # #             color: #1f2937;
# # #         }
# # #         .metric-card {
# # #             background: #ffffff;
# # #             padding: 1.5rem;
# # #             border-radius: 0.75rem;
# # #             box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# # #             text-align: center;
# # #             transition: transform 0.3s ease;
# # #         }
# # #         .metric-card:hover {
# # #             transform: translateY(-5px);
# # #             box-shadow: 0 6px 12px rgba(0,0,0,0.15);
# # #         }
# # #         .metric-label {
# # #             font-size: 1.1rem;
# # #             color: #6b7280;
# # #             margin-bottom: 0.5rem;
# # #         }
# # #         .metric-value {
# # #             font-size: 1.75rem;
# # #             font-weight: 700;
# # #         }
# # #         .stTabs [data-baseweb="tab"] {
# # #             font-size: 1.1rem;
# # #             font-weight: 600;
# # #             padding: 0.75rem 1.5rem;
# # #             border-radius: 0.5rem;
# # #             transition: all 0.3s ease;
# # #         }
# # #         .stTabs [data-baseweb="tab"]:hover {
# # #             background: #e5e7eb;
# # #         }
# # #         .insights-box {
# # #             background: #ffffff;
# # #             padding: 1.5rem;
# # #             border-radius: 0.5rem;
# # #             box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# # #             margin-top: 1.5rem;
# # #         }
# # #         .chart-container {
# # #             background: #ffffff;
# # #             padding: 1.5rem;
# # #             border-radius: 0.75rem;
# # #             box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# # #             margin-bottom: 1.5rem;
# # #         }
# # #         .header-text {
# # #             font-size: 2.25rem;
# # #             font-weight: 800;
# # #             color: #1f2937;
# # #             text-align: center;
# # #             margin-bottom: 1rem;
# # #         }
# # #         .subheader-text {
# # #             font-size: 1.25rem;
# # #             color: #4b5563;
# # #             text-align: center;
# # #             margin-bottom: 2rem;
# # #         }
# # #         footer { visibility: hidden; }
# # #         </style>
# # #     """, unsafe_allow_html=True)

# # #     # Input Section for PNL Dashboard
# # #     st.markdown('<h1 class="header-text">📈 Ultimate Financial PNL Dashboard</h1>', unsafe_allow_html=True)
# # #     st.markdown('<p class="subheader-text">Analyze your portfolio with real-time insights, interactive charts, and comprehensive data exports for smarter trading decisions.</p>', unsafe_allow_html=True)

# # #     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📁 Upload Data</h2>', unsafe_allow_html=True)
# # #     with st.container():
# # #         positions_file = st.file_uploader("Positions CSV", type="csv", help="Upload VS20 22 AUG 2025 POSITIONS(EOD).csv", key="positions_upload")

# # #         # Create two columns for the checkboxes (same row)
# # #         checkbox_col1, checkbox_col2 = st.columns(2)

# # #         with checkbox_col1:
# # #             include_settlement_nfo = st.checkbox(
# # #                 "Include Settlement PNL for NFO",
# # #                 value=True,
# # #                 help="Uncheck to exclude settlement PNL for NFO",
# # #                 key="nfo_settlement"
# # #             )

# # #         with checkbox_col2:
# # #             include_settlement_bfo = st.checkbox(
# # #                 "Include Settlement PNL for BFO",
# # #                 value=True,
# # #                 help="Uncheck to exclude settlement PNL for BFO",
# # #                 key="bfo_settlement"
# # #             )

# # #         # Layout for file uploaders
# # #         col1, col2 = st.columns(2)

# # #         with col1:
# # #             if include_settlement_nfo:
# # #                 nfo_bhav_file = st.file_uploader(
# # #                     "📄 NFO Bhavcopy",
# # #                     type="csv",
# # #                     help="Upload op220825.csv",
# # #                     key="nfo_upload"
# # #                 )
# # #             else:
# # #                 nfo_bhav_file = None
# # #                 st.info("✅ NFO Bhavcopy not required when settlement PNL for NFO is disabled.")

# # #         with col2:
# # #             if include_settlement_bfo:
# # #                 bfo_bhav_file = st.file_uploader(
# # #                     "📄 BFO Bhavcopy",
# # #                     type="csv",
# # #                     help="Upload MS_20250822-01.csv",
# # #                     key="bfo_upload"
# # #                 )
# # #             else:
# # #                 bfo_bhav_file = None
# # #                 st.info("✅ BFO Bhavcopy not required when settlement PNL for BFO is disabled.")

# # #     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📅 Expiry Dates</h2>', unsafe_allow_html=True)
# # #     with st.container():
# # #         col3, col4 = st.columns(2)
# # #         with col3:
# # #             expiry_nfo = st.date_input("NFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="nfo_expiry", disabled=not include_settlement_nfo)
# # #         with col4:
# # #             expiry_bfo = st.date_input("BFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="bfo_expiry", disabled=not include_settlement_bfo)

# # #     # Dark theme adjustments
# # #     if st.session_state.get("theme_select") == "Dark":
# # #         st.markdown("""
# # #             <style>
# # #             body {
# # #                 background: linear-gradient(135deg, #1f2937, #374151);
# # #                 color: #f3f4f6;
# # #             }
# # #             .stDateInput input {
# # #                 background: #4b5563;
# # #                 color: #f3f4f6;
# # #                 border: 2px solid #60a5fa;
# # #             }
# # #             .stDateInput input:focus {
# # #                 border-color: #3b82f6;
# # #                 box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
# # #                 background: #6b7280;
# # #             }
# # #             .metric-card, .insights-box, .chart-container {
# # #                 background: #4b5563;
# # #             }
# # #             .metric-label {
# # #                 color: #d1d5db;
# # #             }
# # #             .header-text, .subheader-text {
# # #                 color: #f3f4f6;
# # #             }
# # #             .stTabs [data-baseweb="tab"] {
# # #                 color: #f3f4f6;
# # #             }
# # #             .stTabs [data-baseweb="tab"]:hover {
# # #                 background: #6b7280;
# # #             }
# # #             </style>
# # #         """, unsafe_allow_html=True)

# # #     process_button = st.button("🚀 Process Data", key="process_button")

# # #     if process_button:
# # #         if positions_file:
# # #             if (include_settlement_nfo and nfo_bhav_file is None) or (include_settlement_bfo and bfo_bhav_file is None):
# # #                 st.error("⚠️ Please upload all required CSV files for enabled settlement calculations.")
# # #             else:
# # #                 try:
# # #                     with st.spinner("🔄 Processing your data..."):
# # #                         results = process_data(positions_file, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo)

# # #                     # Key Metrics Section
# # #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">🔑 Key Financial Metrics</h2>', unsafe_allow_html=True)
# # #                     col1, col2, col3 = st.columns(3, gap="medium")
# # #                     with col1:
# # #                         color = '#34D399' if results["overall_realized"] > 0 else '#F87171' if results["overall_realized"] < 0 else '#9CA3AF'
# # #                         st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_realized"]:,.2f}</p></div>', unsafe_allow_html=True)
# # #                     if include_settlement_nfo or include_settlement_bfo:
# # #                         with col2:
# # #                             color = '#34D399' if results["overall_settlement"] > 0 else '#F87171' if results["overall_settlement"] < 0 else '#9CA3AF'
# # #                             st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_settlement"]:,.2f}</p></div>', unsafe_allow_html=True)
# # #                     with col3:
# # #                         color = '#34D399' if results["grand_total"] > 0 else '#F87171' if results["grand_total"] < 0 else '#9CA3AF'
# # #                         st.markdown(f'<div class="metric-card"><p class="metric-label">Grand Total PNL</p><p class="metric-value" style="color:{color};">₹{results["grand_total"]:,.2f}</p></div>', unsafe_allow_html=True)

# # #                     col4, col5, col6, col7 = st.columns(4, gap="medium")
# # #                     with col4:
# # #                         color = '#34D399' if results["total_realized_nfo"] > 0 else '#F87171' if results["total_realized_nfo"] < 0 else '#9CA3AF'
# # #                         st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# # #                     if include_settlement_nfo:
# # #                         with col5:
# # #                             color = '#34D399' if results["total_settlement_nfo"] > 0 else '#F87171' if results["total_settlement_nfo"] < 0 else '#9CA3AF'
# # #                             st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# # #                     with col6:
# # #                         color = '#34D399' if results["total_realized_bfo"] > 0 else '#F87171' if results["total_realized_bfo"] < 0 else '#9CA3AF'
# # #                         st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# # #                     if include_settlement_bfo:
# # #                         with col7:
# # #                             color = '#34D399' if results["total_settlement_bfo"] > 0 else '#F87171' if results["total_settlement_bfo"] < 0 else '#9CA3AF'
# # #                             st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)

# # #                     # Download Results
# # #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📥 Download Results</h2>', unsafe_allow_html=True)
# # #                     col_download1 = st.columns(1)[0]
# # #                     with col_download1:
# # #                         pnl_breakdown_df = pd.DataFrame({
# # #                             'Date': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
# # #                             'NFO Realized PNL': [results["total_realized_nfo"]],
# # #                             'NFO Settlement PNL': [results["total_settlement_nfo"] if include_settlement_nfo else 0],
# # #                             'Nfo total': [results["total_realized_nfo"] + (results["total_settlement_nfo"] if include_settlement_nfo else 0)],
# # #                             'BFO Realized PNL': [results["total_realized_bfo"]],
# # #                             'BFO Settlement PNL': [results["total_settlement_bfo"] if include_settlement_bfo else 0],
# # #                             'BFO total': [results["total_realized_bfo"] + (results["total_settlement_bfo"] if include_settlement_bfo else 0)],
# # #                             'Grand Total PNL': [results["grand_total"]]
# # #                         })
# # #                         st.markdown(get_excel_download_link(pnl_breakdown_df, "A19_data"), unsafe_allow_html=True)
# # #                         st.success("✅ PNL data processed successfully!")

# # #                 except Exception as e:
# # #                     logger.error(f"Error in process_data: {str(e)}")
# # #                     st.error(f"⚠️ Error processing data: {str(e)}")
# # #                     st.info("Please ensure all uploaded CSV files have the correct columns and data format. Check the log for details.")
# # #         else:
# # #             st.error("⚠️ Please upload the positions CSV file to proceed.")

# # #     # Portfolio Analysis Section
# # #     st.markdown('<hr class="my-8 border-gray-300">', unsafe_allow_html=True)
# # #     st.markdown('<h1 class="header-text">📊 Portfolio Analysis</h1>', unsafe_allow_html=True)
# # #     st.markdown('<p class="subheader-text">Upload GridLog and Summary files to analyze portfolio reasons and timestamps.</p>', unsafe_allow_html=True)
# # #     st.info("Upload the required files below and click 'Process Portfolio Data' to view results.")

# # #     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📁 Upload Portfolio Data</h2>', unsafe_allow_html=True)
# # #     with st.container():
# # #         col_grid, col_summary = st.columns(2)
# # #         with col_grid:
# # #             gridlog_file = st.file_uploader(
# # #                 "GridLog File",
# # #                 type=["csv", "xlsx"],
# # #                 help="Upload VS20 24 OCT 2025 GridLog.csv or .xlsx",
# # #                 key="gridlog_upload"
# # #             )
# # #         with col_summary:
# # #             summary_file = st.file_uploader(
# # #                 "Summary Excel File",
# # #                 type="xlsx",
# # #                 help="Upload VS20 24 OCT 2025 SUMMARY.xlsx",
# # #                 key="summary_upload"
# # #             )

# # #     process_portfolio_button = st.button("🚀 Process Portfolio Data", key="process_portfolio_button")

# # #     if process_portfolio_button:
# # #         if gridlog_file and summary_file:
# # #             try:
# # #                 with st.spinner("🔄 Processing portfolio data..."):
# # #                     logger.info("Starting portfolio data processing")
# # #                     final_df = process_portfolio_data(gridlog_file, summary_file)

# # #                     # Display Results
# # #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📋 Portfolio Analysis Results</h2>', unsafe_allow_html=True)
# # #                     st.markdown('<div class="insights-box">', unsafe_allow_html=True)
# # #                     st.write(final_df)
# # #                     st.markdown('</div>', unsafe_allow_html=True)
# # #                     st.success("✅ Portfolio data processed successfully!")

# # #                     # Download Link
# # #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📥 Download Portfolio Results</h2>', unsafe_allow_html=True)
# # #                     output_path = "final_portfolios_with_reason_and_time_24_oct"
# # #                     st.markdown(get_csv_download_link(final_df, output_path), unsafe_allow_html=True)

# # #             except Exception as e:
# # #                 logger.error(f"Error in process_portfolio_data: {str(e)}")
# # #                 st.error(f"⚠️ Error processing portfolio data: {str(e)}")
# # #                 st.info("Please ensure the uploaded files have the correct format and columns (e.g., 'Message', 'Option Portfolio', 'Timestamp' for GridLog; 'Exit Type', 'Portfolio Name', 'Exit Time', 'Status' for Summary).")
# # #         else:
# # #             st.error("⚠️ Please upload both GridLog and Summary files to proceed.")

# # # # Function to process portfolio data
# # # def process_portfolio_data(gridlog_file, summary_file):
# # #     logger.info("Loading GridLog file")
# # #     if gridlog_file.name.endswith('.csv'):
# # #         df_grid = pd.read_csv(gridlog_file)
# # #     elif gridlog_file.name.endswith('.xlsx'):
# # #         df_grid = pd.read_excel(gridlog_file)
# # #     else:
# # #         raise ValueError("Unsupported GridLog file type. Use CSV or Excel.")

# # #     logger.info("Cleaning GridLog column names")
# # #     df_grid.columns = df_grid.columns.str.strip()

# # #     logger.info("Filtering GridLog for Combined SL/Trail Target messages")
# # #     mask = df_grid['Message'].str.contains(r'Combined SL:|Combined trail target:', case=False, na=False)
# # #     filtered_grid = df_grid.loc[mask, ['Message', 'Option Portfolio', 'Timestamp']].dropna(subset=['Option Portfolio'])

# # #     logger.info("Grouping GridLog results")
# # #     summary_grid = (
# # #         filtered_grid.groupby('Option Portfolio').agg({
# # #             'Message': lambda x: ', '.join(x.unique()),
# # #             'Timestamp': 'max'
# # #         }).reset_index()
# # #         .rename(columns={'Message': 'Reason', 'Timestamp': 'Time'})
# # #     )

# # #     logger.info("Processing Summary Excel file")
# # #     xl = pd.ExcelFile(summary_file)
# # #     summary_list = []

# # #     for sheet_name in xl.sheet_names:
# # #         if "legs" in sheet_name.lower():
# # #             df_leg = xl.parse(sheet_name)
# # #             df_leg.columns = df_leg.columns.str.strip()

# # #             if {'Exit Type', 'Portfolio Name', 'Exit Time'}.issubset(df_leg.columns):
# # #                 onsqoff_df = df_leg[df_leg['Exit Type'].astype(str).str.strip() == 'OnSqOffTime']
# # #                 if not onsqoff_df.empty:
# # #                     grouped = onsqoff_df.groupby('Portfolio Name')['Exit Time'].max().reset_index()
# # #                     for _, row in grouped.iterrows():
# # #                         summary_list.append({
# # #                             'Option Portfolio': row['Portfolio Name'],
# # #                             'Reason': 'OnSqOffTime',
# # #                             'Time': row['Exit Time']
# # #                         })

# # #     summary_summary = pd.DataFrame(summary_list)

# # #     logger.info("Combining GridLog and Summary results")
# # #     final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)
# # #     final_df = (
# # #         final_df.groupby('Option Portfolio').agg({
# # #             'Reason': lambda x: ', '.join(sorted(set(x))),
# # #             'Time': 'last'
# # #         }).reset_index()
# # #     )

# # #     logger.info("Adding completed portfolios")
# # #     completed_list = []
# # #     grid_portfolios = df_grid['Option Portfolio'].dropna().unique()

# # #     for sheet_name in xl.sheet_names:
# # #         if "legs" in sheet_name.lower():
# # #             df_leg = xl.parse(sheet_name)
# # #             df_leg.columns = df_leg.columns.str.strip()

# # #             if 'Portfolio Name' in df_leg.columns and 'Status' in df_leg.columns:
# # #                 for portfolio, group in df_leg.groupby('Portfolio Name'):
# # #                     if portfolio not in final_df['Option Portfolio'].values and portfolio in grid_portfolios:
# # #                         statuses = group['Status'].astype(str).str.strip().unique()
# # #                         if len(statuses) == 1 and statuses[0].lower() == 'completed':
# # #                             reason_text = 'AllLegsCompleted'
# # #                             exit_time_to_use = None

# # #                             if 'Exit Time' in group.columns:
# # #                                 for exit_time, exit_type in zip(group['Exit Time'], group.get('Exit Type', [])):
# # #                                     if pd.isna(exit_time):
# # #                                         continue
# # #                                     normalized_exit_time = str(exit_time).replace('.', ':').strip()
# # #                                     matching_rows = df_grid[
# # #                                         (df_grid['Option Portfolio'] == portfolio) &
# # #                                         (df_grid['Timestamp'].astype(str).str.contains(normalized_exit_time))
# # #                                     ]
# # #                                     if not matching_rows.empty:
# # #                                         reason_text += f", {exit_type.strip()}"
# # #                                         exit_time_to_use = exit_time
# # #                                         break

# # #                             completed_list.append({
# # #                                 'Option Portfolio': portfolio,
# # #                                 'Reason': reason_text,
# # #                                 'Time': exit_time_to_use
# # #                             })

# # #     if completed_list:
# # #         completed_df = pd.DataFrame(completed_list)
# # #         final_df = pd.concat([final_df, completed_df], ignore_index=True)

# # #     logger.info("Cleaning Reason texts")
# # #     def clean_reason(text):
# # #         if pd.isna(text):
# # #             return text
# # #         text = str(text)
# # #         match = re.search(r'(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)', text, re.IGNORECASE)
# # #         if match:
# # #             return match.group(1)
# # #         if 'AllLegsCompleted' in text:
# # #             text = text.replace('AllLegsCompleted,', '').strip()
# # #             text = text.replace('AllLegsCompleted', '').strip()
# # #         return text.strip()

# # #     final_df['Reason'] = final_df['Reason'].apply(clean_reason)

# # #     logger.info("Cleaning Time column")
# # #     final_df['Time'] = final_df['Time'].astype(str).str.strip().replace('nan', None)

# # #     logger.info("Portfolio data processing completed successfully")
# # #     return final_df

# # # # Function to process PNL data
# # # def process_data(positions_file, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo):
# # #     logger.info("Starting PNL data processing")
# # #     try:
# # #         df = pd.read_csv(positions_file)
# # #         logger.info("Positions CSV loaded successfully")

# # #         required_columns = ['Exchange', 'Symbol', 'Net Qty', 'Buy Avg Price', 'Sell Avg Price', 'Sell Qty', 'Buy Qty', 'Realized Profit', 'Unrealized Profit']
# # #         missing_columns = [col for col in required_columns if col not in df.columns]
# # #         if missing_columns:
# # #             raise ValueError(f"Missing columns in positions file: {missing_columns}")

# # #         mask = df["Exchange"].isin(["NFO", "BFO"])
# # #         if 'Symbol' in df.columns:
# # #             df.loc[mask, "Symbol"] = df.loc[mask, "Symbol"].astype(str).str[-5:] + df.loc[mask, "Symbol"].astype(str).str[-8:-6]
# # #         else:
# # #             raise KeyError("Symbol column not found in positions file")

# # #         df_nfo = df[df["Exchange"] == "NFO"].copy()
# # #         df_bfo = df[df["Exchange"] == "BFO"].copy()
# # #         logger.info("Data split into NFO and BFO")

# # #         total_settlement_pnl_nfo = 0
# # #         total_settlement_pnl_bfo = 0
# # #         df_bhav_nfo = pd.DataFrame()
# # #         df_bhav_bfo = pd.DataFrame()

# # #         if include_settlement_nfo:
# # #             if nfo_bhav_file is None:
# # #                 raise ValueError("NFO bhavcopy file is required when settlement PNL for NFO is enabled")
# # #             df_bhav_nfo = pd.read_csv(nfo_bhav_file)
# # #             logger.info("NFO bhavcopy loaded successfully")

# # #             if 'CONTRACT_D' not in df_bhav_nfo.columns:
# # #                 raise ValueError("CONTRACT_D column not found in NFO bhavcopy")
# # #             df_bhav_nfo["Date"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(\d{2}-[A-Z]{3}-\d{4})')
# # #             df_bhav_nfo["Symbol"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'^(.*?)(\d{2}-[A-Z]{3}-\d{4})')[0]
# # #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(PE\d+|CE\d+)$')
# # #             df_bhav_nfo["Date"] = pd.to_datetime(df_bhav_nfo["Date"], format="%d-%b-%Y")
# # #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["Strike_Type"].str.replace(
# # #                 r'^(PE|CE)(\d+)$', r'\2\1', regex=True
# # #             )

# # #             target_symbol = "OPTIDXNIFTY"
# # #             df_bhav_nfo = df_bhav_nfo[
# # #                 (df_bhav_nfo["Date"] == pd.to_datetime(expiry_nfo)) &
# # #                 (df_bhav_nfo["Symbol"] == target_symbol)
# # #             ]

# # #             df_nfo["Strike_Type"] = df_nfo["Symbol"].str.extract(r'(\d+[A-Z]{2})$')
# # #             df_nfo['Strike'] = df_nfo['Strike_Type'].str[:-2].astype(float, errors='ignore')
# # #             df_nfo['Option_Type'] = df_nfo['Strike_Type'].str[-2:]

# # #             df_nfo = df_nfo.merge(
# # #                 df_bhav_nfo[["Strike_Type", "SETTLEMENT"]],
# # #                 on="Strike_Type",
# # #                 how="left"
# # #             )
# # #             logger.info("NFO data merged with bhavcopy")

# # #         if include_settlement_bfo:
# # #             if bfo_bhav_file is None:
# # #                 raise ValueError("BFO bhavcopy file is required when settlement PNL for BFO is enabled")
# # #             df_bhav_bfo = pd.read_csv(bfo_bhav_file)
# # #             logger.info("BFO bhavcopy loaded successfully")

# # #             if 'Market Summary Date' not in df_bhav_bfo.columns or 'Expiry Date' not in df_bhav_bfo.columns or 'Series Code' not in df_bhav_bfo.columns:
# # #                 raise ValueError("Required columns missing in BFO bhavcopy")
# # #             df_bhav_bfo["Date"] = pd.to_datetime(df_bhav_bfo["Market Summary Date"], format="%d %b %Y", errors="coerce")
# # #             df_bhav_bfo["Expiry Date"] = pd.to_datetime(df_bhav_bfo["Expiry Date"], format="%d %b %Y", errors="coerce")
# # #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Series Code"].astype(str).str[-7:]

# # #             df_bhav_bfo = df_bhav_bfo[
# # #                 (df_bhav_bfo["Expiry Date"] == pd.to_datetime(expiry_bfo))
# # #             ]

# # #             df_bfo["Symbol"] = df_bfo["Symbol"].astype(str).str.strip()
# # #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Symbols"].astype(str).str.strip()

# # #             bhav_mapping = df_bhav_bfo.drop_duplicates(subset="Symbols", keep="last").set_index("Symbols")["Close Price"]
# # #             df_bfo["Close Price"] = df_bfo["Symbol"].map(bhav_mapping)
# # #             logger.info("BFO data processed")

# # #         logger.info("Calculating Realized PNL for NFO")
# # #         conditions = [
# # #             df_nfo["Net Qty"] == 0,
# # #             df_nfo["Net Qty"] > 0,
# # #             df_nfo["Net Qty"] < 0
# # #         ]
# # #         choices = [
# # #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# # #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# # #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Buy Qty"]
# # #         ]
# # #         df_nfo["Calculated_Realized_PNL"] = np.select(conditions, choices, default=0)
# # #         df_nfo["Matching_Realized"] = df_nfo["Realized Profit"] == df_nfo["Calculated_Realized_PNL"]
# # #         df_nfo["Matching_Realized"] = df_nfo["Matching_Realized"].replace({True: "TRUE", False: ""})

# # #         if include_settlement_nfo:
# # #             logger.info("Calculating Settlement PNL for NFO")
# # #             df_nfo["Calculated_Settlement_PNL"] = np.select(
# # #                 [
# # #                     df_nfo["Net Qty"] > 0,
# # #                     df_nfo["Net Qty"] < 0
# # #                 ],
# # #                 [
# # #                     (df_nfo["SETTLEMENT"] - df_nfo["Buy Avg Price"]) * abs(df_nfo["Net Qty"]),
# # #                     (df_nfo["Sell Avg Price"] - df_nfo["SETTLEMENT"]) * abs(df_nfo["Net Qty"])
# # #                 ],
# # #                 default=0
# # #             )
# # #             df_nfo["Matching_Settlement"] = df_nfo["Unrealized Profit"] == df_nfo["Calculated_Settlement_PNL"]
# # #             df_nfo["Matching_Settlement"] = df_nfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# # #             total_settlement_pnl_nfo = df_nfo["Calculated_Settlement_PNL"].fillna(0).sum()
# # #         else:
# # #             df_nfo["Calculated_Settlement_PNL"] = 0
# # #             df_nfo["Matching_Settlement"] = ""

# # #         total_realized_pnl_nfo = df_nfo["Calculated_Realized_PNL"].fillna(0).sum()

# # #         logger.info("Calculating Realized PNL for BFO")
# # #         conditions_bfo = [
# # #             df_bfo["Net Qty"] == 0,
# # #             df_bfo["Net Qty"] > 0,
# # #             df_bfo["Net Qty"] < 0
# # #         ]
# # #         choices_bfo = [
# # #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# # #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# # #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Buy Qty"]
# # #         ]
# # #         df_bfo["Calculated_Realized_PNL"] = np.select(conditions_bfo, choices_bfo, default=0)
# # #         df_bfo["Matching_Realized"] = df_bfo["Realized Profit"] == df_bfo["Calculated_Realized_PNL"]
# # #         df_bfo["Matching_Realized"] = df_bfo["Matching_Realized"].replace({True: "TRUE", False: ""})

# # #         if include_settlement_bfo:
# # #             logger.info("Calculating Settlement PNL for BFO")
# # #             long_condition = df_bfo["Net Qty"] > 0
# # #             df_bfo.loc[long_condition, "Calculated_Settlement_PNL"] = (
# # #                 (df_bfo["Close Price"] - df_bfo["Buy Avg Price"]) * abs(df_bfo["Net Qty"])
# # #             )
# # #             short_condition = df_bfo["Net Qty"] < 0
# # #             df_bfo.loc[short_condition, "Calculated_Settlement_PNL"] = (
# # #                 (df_bfo["Sell Avg Price"] - df_bfo["Close Price"]) * abs(df_bfo["Net Qty"])
# # #             )
# # #             df_bfo.loc[df_bfo["Net Qty"] == 0, "Calculated_Settlement_PNL"] = 0
# # #             df_bfo["Matching_Settlement"] = df_bfo["Unrealized Profit"] == df_bfo["Calculated_Settlement_PNL"]
# # #             df_bfo["Matching_Settlement"] = df_bfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# # #             total_settlement_pnl_bfo = df_bfo["Calculated_Settlement_PNL"].fillna(0).sum()
# # #         else:
# # #             df_bfo["Calculated_Settlement_PNL"] = 0
# # #             df_bfo["Matching_Settlement"] = ""

# # #         total_realized_pnl_bfo = df_bfo["Calculated_Realized_PNL"].fillna(0).sum()

# # #         overall_realized = total_realized_pnl_nfo + total_realized_pnl_bfo
# # #         overall_settlement = total_settlement_pnl_nfo + total_settlement_pnl_bfo
# # #         grand_total = overall_realized + overall_settlement

# # #         logger.info("PNL data processing completed successfully")
# # #         return {
# # #             "df_nfo": df_nfo,
# # #             "df_bfo": df_bfo,
# # #             "total_realized_nfo": total_realized_pnl_nfo,
# # #             "total_settlement_nfo": total_settlement_pnl_nfo,
# # #             "total_realized_bfo": total_realized_pnl_bfo,
# # #             "total_settlement_bfo": total_settlement_pnl_bfo,
# # #             "overall_realized": overall_realized,
# # #             "overall_settlement": overall_settlement,
# # #             "grand_total": grand_total
# # #         }
# # #     except Exception as e:
# # #         logger.error(f"Error in process_data: {str(e)}")
# # #         raise

# # # # Function to get Excel download link
# # # def get_excel_download_link(df, filename):
# # #     output = BytesIO()
# # #     with pd.ExcelWriter(output, engine='openpyxl') as writer:
# # #         df.to_excel(writer, index=False, sheet_name='PNL Data')
# # #         worksheet = writer.sheets['PNL Data']
# # #         for row in worksheet.rows:
# # #             for cell in row:
# # #                 cell.alignment = Alignment(horizontal='center')
# # #         for cell in worksheet[1]:
# # #             cell.font = Font(bold=True, color="FFFFFF")
# # #             cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
# # #         for row in worksheet.rows:
# # #             for cell in row:
# # #                 cell.border = Border(left=Side(style='thin'),
# # #                                     right=Side(style='thin'),
# # #                                     top=Side(style='thin'),
# # #                                     bottom=Side(style='thin'))
# # #     excel_data = output.getvalue()
# # #     b64 = base64.b64encode(excel_data).decode()
# # #     return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}.xlsx</a>'

# # # # Function to get CSV download link
# # # def get_csv_download_link(df, filename):
# # #     output = BytesIO()
# # #     df.to_csv(output, index=False)
# # #     csv_data = output.getvalue()
# # #     b64 = base64.b64encode(csv_data).decode()
# # #     return f'<a href="data:text/csv;base64,{b64}" download="{filename}.csv" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}.csv</a>'

# # # # Run the app
# # # if __name__ == "__main__":
# # #     st.write(f"DEBUG: Starting Streamlit app at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
# # #     run()
# # ########################################################################################################################################################
# # import streamlit as st
# # import pandas as pd
# # import numpy as np
# # import re
# # import base64
# # from io import BytesIO
# # import logging
# # import openpyxl
# # from openpyxl.styles import Alignment, Border, Side, Font
# # from datetime import datetime
# # # Set up logging
# # logging.basicConfig(level=logging.INFO)
# # logger = logging.getLogger(__name__)
# # # Check Streamlit version for compatibility
# # try:
# #     import streamlit
# #     logger.info(f"Streamlit version: {streamlit.__version__}")
# # except ImportError:
# #     st.error("Streamlit is not installed. Please install it using `pip install streamlit`.")
# #     st.stop()
# # def run():
# #     # Main title and description
# #     st.markdown("# Realized PNL for 19")
# #     st.write("This tool displays realized profit and loss for 19.")
# #     # Enhanced Custom CSS with Tailwind CSS and Animations
# #     st.markdown("""
# #     <link href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css" rel="stylesheet">
# #     <style>
# #     body {
# #         font-family: 'Inter', sans-serif;
# #         background: linear-gradient(135deg, #f3f4f6, #e5e7eb);
# #     }
# #     .stButton>button {
# #         background: linear-gradient(45deg, #3b82f6, #60a5fa);
# #         color: white;
# #         border: none;
# #         padding: 0.75rem 1.5rem;
# #         border-radius: 0.5rem;
# #         font-weight: 600;
# #         transition: all 0.3s ease;
# #         width: 100%;
# #     }
# #     .stButton>button:hover {
# #         background: linear-gradient(45deg, #2563eb, #3b82f6);
# #         transform: translateY(-2px);
# #         box-shadow: 0 4px 6px rgba(0,0,0,0.2);
# #     }
# #     .stDateInput input {
# #         border: 2px solid #3b82f6;
# #         border-radius: 0.5rem;
# #         padding: 0.5rem;
# #         background: #ffffff;
# #         color: #1f2937;
# #         font-size: 1rem;
# #         transition: all 0.3s ease;
# #     }
# #     .stDateInput input:focus {
# #         outline: none;
# #         border-color: #2563eb;
# #         box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
# #         background: #f8fafc;
# #     }
# #     .stFileUploader button {
# #         background: linear-gradient(45deg, #10b981, #34d399);
# #         color: white;
# #         border-radius: 0.5rem;
# #         padding: 0.75rem;
# #     }
# #     .stFileUploader button:hover {
# #         background: linear-gradient(45deg, #059669, #10b981);
# #     }
# #     .stCheckbox label {
# #         font-size: 1rem;
# #         color: #1f2937;
# #     }
# #     .metric-card {
# #         background: #ffffff;
# #         padding: 1.5rem;
# #         border-radius: 0.75rem;
# #         box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# #         text-align: center;
# #         transition: transform 0.3s ease;
# #     }
# #     .metric-card:hover {
# #         transform: translateY(-5px);
# #         box-shadow: 0 6px 12px rgba(0,0,0,0.15);
# #     }
# #     .metric-label {
# #         font-size: 1.1rem;
# #         color: #6b7280;
# #         margin-bottom: 0.5rem;
# #     }
# #     .metric-value {
# #         font-size: 1.75rem;
# #         font-weight: 700;
# #     }
# #     .stTabs [data-baseweb="tab"] {
# #         font-size: 1.1rem;
# #         font-weight: 600;
# #         padding: 0.75rem 1.5rem;
# #         border-radius: 0.5rem;
# #         transition: all 0.3s ease;
# #     }
# #     .stTabs [data-baseweb="tab"]:hover {
# #         background: #e5e7eb;
# #     }
# #     .insights-box {
# #         background: #ffffff;
# #         padding: 1.5rem;
# #         border-radius: 0.5rem;
# #         box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# #         margin-top: 1.5rem;
# #     }
# #     .chart-container {
# #         background: #ffffff;
# #         padding: 1.5rem;
# #         border-radius: 0.75rem;
# #         box-shadow: 0 4px 6px rgba(0,0,0,0.1);
# #         margin-bottom: 1.5rem;
# #     }
# #     .header-text {
# #         font-size: 2.25rem;
# #         font-weight: 800;
# #         color: #1f2937;
# #         text-align: center;
# #         margin-bottom: 1rem;
# #     }
# #     .subheader-text {
# #         font-size: 1.25rem;
# #         color: #4b5563;
# #         text-align: center;
# #         margin-bottom: 2rem;
# #     }
# #     footer { visibility: hidden; }
# #     </style>
# #     """, unsafe_allow_html=True)
# #     # Input Section for PNL Dashboard
# #     st.markdown('<h1 class="header-text">📈 Ultimate Financial PNL Dashboard</h1>', unsafe_allow_html=True)
# #     st.markdown('<p class="subheader-text">Analyze your portfolio with real-time insights, interactive charts, and comprehensive data exports for smarter trading decisions.</p>', unsafe_allow_html=True)
# #     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📁 Upload Data</h2>', unsafe_allow_html=True)
# #     with st.container():
# #         positions_file = st.file_uploader("Positions CSV", type="csv", help="Upload VS20 22 AUG 2025 POSITIONS(EOD).csv", key="positions_upload")
# #         selected_user = None
# #         if positions_file:
# #             if 'positions_df' not in st.session_state or st.session_state.get('positions_file_name') != positions_file.name:
# #                 st.session_state.positions_df = pd.read_csv(positions_file)
# #                 st.session_state.positions_file_name = positions_file.name
# #             df = st.session_state.positions_df
# #             if 'UserID' in df.columns:
# #                 users = sorted(df['UserID'].unique().tolist())
# #                 selected_user = st.selectbox("Select User", users, key="selected_user")
# #             else:
# #                 st.error("⚠️ 'UserID' column not found in positions file.")
# #         # Create two columns for the checkboxes (same row)
# #         checkbox_col1, checkbox_col2 = st.columns(2)
# #         with checkbox_col1:
# #             include_settlement_nfo = st.checkbox(
# #                 "Include Settlement PNL for NFO",
# #                 value=True,
# #                 help="Uncheck to exclude settlement PNL for NFO",
# #                 key="nfo_settlement"
# #             )
# #         with checkbox_col2:
# #             include_settlement_bfo = st.checkbox(
# #                 "Include Settlement PNL for BFO",
# #                 value=True,
# #                 help="Uncheck to exclude settlement PNL for BFO",
# #                 key="bfo_settlement"
# #             )
# #         # Layout for file uploaders
# #         col1, col2 = st.columns(2)
# #         with col1:
# #             if include_settlement_nfo:
# #                 nfo_bhav_file = st.file_uploader(
# #                     "📄 NFO Bhavcopy",
# #                     type="csv",
# #                     help="Upload op220825.csv",
# #                     key="nfo_upload"
# #                 )
# #             else:
# #                 nfo_bhav_file = None
# #                 st.info("✅ NFO Bhavcopy not required when settlement PNL for NFO is disabled.")
# #         with col2:
# #             if include_settlement_bfo:
# #                 bfo_bhav_file = st.file_uploader(
# #                     "📄 BFO Bhavcopy",
# #                     type="csv",
# #                     help="Upload MS_20250822-01.csv",
# #                     key="bfo_upload"
# #                 )
# #             else:
# #                 bfo_bhav_file = None
# #                 st.info("✅ BFO Bhavcopy not required when settlement PNL for BFO is disabled.")
# #         st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📅 Expiry Dates</h2>', unsafe_allow_html=True)
# #         with st.container():
# #             col3, col4 = st.columns(2)
# #             with col3:
# #                 expiry_nfo = st.date_input("NFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="nfo_expiry", disabled=not include_settlement_nfo)
# #             with col4:
# #                 expiry_bfo = st.date_input("BFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="bfo_expiry", disabled=not include_settlement_bfo)
# #         # Dark theme adjustments
# #         if st.session_state.get("theme_select") == "Dark":
# #             st.markdown("""
# #             <style>
# #             body {
# #                 background: linear-gradient(135deg, #1f2937, #374151);
# #                 color: #f3f4f6;
# #             }
# #             .stDateInput input {
# #                 background: #4b5563;
# #                 color: #f3f4f6;
# #                 border: 2px solid #60a5fa;
# #             }
# #             .stDateInput input:focus {
# #                 border-color: #3b82f6;
# #                 box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
# #                 background: #6b7280;
# #             }
# #             .metric-card, .insights-box, .chart-container {
# #                 background: #4b5563;
# #             }
# #             .metric-label {
# #                 color: #d1d5db;
# #             }
# #             .header-text, .subheader-text {
# #                 color: #f3f4f6;
# #             }
# #             .stTabs [data-baseweb="tab"] {
# #                 color: #f3f4f6;
# #             }
# #             .stTabs [data-baseweb="tab"]:hover {
# #                 background: #6b7280;
# #             }
# #             </style>
# #             """, unsafe_allow_html=True)
# #         process_button = st.button("🚀 Process Data", key="process_button")
# #         if process_button:
# #             if positions_file and selected_user:
# #                 if (include_settlement_nfo and nfo_bhav_file is None) or (include_settlement_bfo and bfo_bhav_file is None):
# #                     st.error("⚠️ Please upload all required CSV files for enabled settlement calculations.")
# #                 else:
# #                     try:
# #                         with st.spinner("🔄 Processing your data..."):
# #                             filtered_df = st.session_state.positions_df[st.session_state.positions_df['UserID'] == selected_user]
# #                             results = process_data(filtered_df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo)
# #                         # Key Metrics Section
# #                         st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">🔑 Key Financial Metrics</h2>', unsafe_allow_html=True)
# #                         col1, col2, col3 = st.columns(3, gap="medium")
# #                         with col1:
# #                             color = '#34D399' if results["overall_realized"] > 0 else '#F87171' if results["overall_realized"] < 0 else '#9CA3AF'
# #                             st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_realized"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         if include_settlement_nfo or include_settlement_bfo:
# #                             with col2:
# #                                 color = '#34D399' if results["overall_settlement"] > 0 else '#F87171' if results["overall_settlement"] < 0 else '#9CA3AF'
# #                                 st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_settlement"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                             with col3:
# #                                 color = '#34D399' if results["grand_total"] > 0 else '#F87171' if results["grand_total"] < 0 else '#9CA3AF'
# #                                 st.markdown(f'<div class="metric-card"><p class="metric-label">Grand Total PNL</p><p class="metric-value" style="color:{color};">₹{results["grand_total"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         col4, col5, col6, col7 = st.columns(4, gap="medium")
# #                         with col4:
# #                             color = '#34D399' if results["total_realized_nfo"] > 0 else '#F87171' if results["total_realized_nfo"] < 0 else '#9CA3AF'
# #                             st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         if include_settlement_nfo:
# #                             with col5:
# #                                 color = '#34D399' if results["total_settlement_nfo"] > 0 else '#F87171' if results["total_settlement_nfo"] < 0 else '#9CA3AF'
# #                                 st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         with col6:
# #                             color = '#34D399' if results["total_realized_bfo"] > 0 else '#F87171' if results["total_realized_bfo"] < 0 else '#9CA3AF'
# #                             st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         if include_settlement_bfo:
# #                             with col7:
# #                                 color = '#34D399' if results["total_settlement_bfo"] > 0 else '#F87171' if results["total_settlement_bfo"] < 0 else '#9CA3AF'
# #                                 st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)
# #                         # Download Results
# #                         st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📥 Download Results</h2>', unsafe_allow_html=True)
# #                         col_download1 = st.columns(1)[0]
# #                         with col_download1:
# #                             pnl_breakdown_df = pd.DataFrame({
# #                                 'Date': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
# #                                 'NFO Realized PNL': [results["total_realized_nfo"]],
# #                                 'NFO Settlement PNL': [results["total_settlement_nfo"] if include_settlement_nfo else 0],
# #                                 'Nfo total': [results["total_realized_nfo"] + (results["total_settlement_nfo"] if include_settlement_nfo else 0)],
# #                                 'BFO Realized PNL': [results["total_realized_bfo"]],
# #                                 'BFO Settlement PNL': [results["total_settlement_bfo"] if include_settlement_bfo else 0],
# #                                 'BFO total': [results["total_realized_bfo"] + (results["total_settlement_bfo"] if include_settlement_bfo else 0)],
# #                                 'Grand Total PNL': [results["grand_total"]]
# #                             })
# #                             st.markdown(get_excel_download_link(pnl_breakdown_df, "A19_data"), unsafe_allow_html=True)
# #                         st.success("✅ PNL data processed successfully!")
# #                     except Exception as e:
# #                         logger.error(f"Error in process_data: {str(e)}")
# #                         st.error(f"⚠️ Error processing data: {str(e)}")
# #                         st.info("Please ensure all uploaded CSV files have the correct columns and data format. Check the log for details.")
# #             else:
# #                 st.error("⚠️ Please upload the positions CSV file and select a user to proceed.")
# #     # Portfolio Analysis Section
# #     st.markdown('<hr class="my-8 border-gray-300">', unsafe_allow_html=True)
# #     st.markdown('<h1 class="header-text">📊 Portfolio Analysis</h1>', unsafe_allow_html=True)
# #     st.markdown('<p class="subheader-text">Upload GridLog and Summary files to analyze portfolio reasons and timestamps.</p>', unsafe_allow_html=True)
# #     st.info("Upload the required files below and click 'Process Portfolio Data' to view results.")
# #     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">📁 Upload Portfolio Data</h2>', unsafe_allow_html=True)
# #     with st.container():
# #         col_grid, col_summary = st.columns(2)
# #         with col_grid:
# #             gridlog_file = st.file_uploader(
# #                 "GridLog File",
# #                 type=["csv", "xlsx"],
# #                 help="Upload VS20 24 OCT 2025 GridLog.csv or .xlsx",
# #                 key="gridlog_upload"
# #             )
# #         with col_summary:
# #             summary_file = st.file_uploader(
# #                 "Summary Excel File",
# #                 type="xlsx",
# #                 help="Upload VS20 24 OCT 2025 SUMMARY.xlsx",
# #                 key="summary_upload"
# #             )
# #         process_portfolio_button = st.button("🚀 Process Portfolio Data", key="process_portfolio_button")
# #         if process_portfolio_button:
# #             if gridlog_file and summary_file:
# #                 try:
# #                     with st.spinner("🔄 Processing portfolio data..."):
# #                         logger.info("Starting portfolio data processing")
# #                         final_df = process_portfolio_data(gridlog_file, summary_file)
# #                     # Display Results
# #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📋 Portfolio Analysis Results</h2>', unsafe_allow_html=True)
# #                     st.markdown('<div class="insights-box">', unsafe_allow_html=True)
# #                     st.write(final_df)
# #                     st.markdown('</div>', unsafe_allow_html=True)
# #                     st.success("✅ Portfolio data processed successfully!")
# #                     # Download Link
# #                     st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">📥 Download Portfolio Results</h2>', unsafe_allow_html=True)
# #                     output_path = "final_portfolios_with_reason_and_time_24_oct"
# #                     st.markdown(get_csv_download_link(final_df, output_path), unsafe_allow_html=True)
# #                 except Exception as e:
# #                     logger.error(f"Error in process_portfolio_data: {str(e)}")
# #                     st.error(f"⚠️ Error processing portfolio data: {str(e)}")
# #                     st.info("Please ensure the uploaded files have the correct format and columns (e.g., 'Message', 'Option Portfolio', 'Timestamp' for GridLog; 'Exit Type', 'Portfolio Name', 'Exit Time', 'Status' for Summary).")
# #             else:
# #                 st.error("⚠️ Please upload both GridLog and Summary files to proceed.")
# # # Function to process portfolio data
# # def process_portfolio_data(gridlog_file, summary_file):
# #     logger.info("Loading GridLog file")
# #     if gridlog_file.name.endswith('.csv'):
# #         df_grid = pd.read_csv(gridlog_file)
# #     elif gridlog_file.name.endswith('.xlsx'):
# #         df_grid = pd.read_excel(gridlog_file)
# #     else:
# #         raise ValueError("Unsupported GridLog file type. Use CSV or Excel.")
# #     logger.info("Cleaning GridLog column names")
# #     df_grid.columns = df_grid.columns.str.strip()
# #     logger.info("Filtering GridLog for Combined SL/Trail Target messages")
# #     mask = df_grid['Message'].str.contains(r'Combined SL:|Combined trail target:', case=False, na=False)
# #     filtered_grid = df_grid.loc[mask, ['Message', 'Option Portfolio', 'Timestamp']].dropna(subset=['Option Portfolio'])
# #     logger.info("Grouping GridLog results")
# #     summary_grid = (
# #         filtered_grid.groupby('Option Portfolio').agg({
# #             'Message': lambda x: ', '.join(x.unique()),
# #             'Timestamp': 'max'
# #         }).reset_index()
# #         .rename(columns={'Message': 'Reason', 'Timestamp': 'Time'})
# #     )
# #     logger.info("Processing Summary Excel file")
# #     xl = pd.ExcelFile(summary_file)
# #     summary_list = []
# #     for sheet_name in xl.sheet_names:
# #         if "legs" in sheet_name.lower():
# #             df_leg = xl.parse(sheet_name)
# #             df_leg.columns = df_leg.columns.str.strip()
# #             if {'Exit Type', 'Portfolio Name', 'Exit Time'}.issubset(df_leg.columns):
# #                 onsqoff_df = df_leg[df_leg['Exit Type'].astype(str).str.strip() == 'OnSqOffTime']
# #                 if not onsqoff_df.empty:
# #                     grouped = onsqoff_df.groupby('Portfolio Name')['Exit Time'].max().reset_index()
# #                     for _, row in grouped.iterrows():
# #                         summary_list.append({
# #                             'Option Portfolio': row['Portfolio Name'],
# #                             'Reason': 'OnSqOffTime',
# #                             'Time': row['Exit Time']
# #                         })
# #     summary_summary = pd.DataFrame(summary_list)
# #     logger.info("Combining GridLog and Summary results")
# #     final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)
# #     final_df = (
# #         final_df.groupby('Option Portfolio').agg({
# #             'Reason': lambda x: ', '.join(sorted(set(x))),
# #             'Time': 'last'
# #         }).reset_index()
# #     )
# #     logger.info("Adding completed portfolios")
# #     completed_list = []
# #     grid_portfolios = df_grid['Option Portfolio'].dropna().unique()
# #     for sheet_name in xl.sheet_names:
# #         if "legs" in sheet_name.lower():
# #             df_leg = xl.parse(sheet_name)
# #             df_leg.columns = df_leg.columns.str.strip()
# #             if 'Portfolio Name' in df_leg.columns and 'Status' in df_leg.columns:
# #                 for portfolio, group in df_leg.groupby('Portfolio Name'):
# #                     if portfolio not in final_df['Option Portfolio'].values and portfolio in grid_portfolios:
# #                         statuses = group['Status'].astype(str).str.strip().unique()
# #                         if len(statuses) == 1 and statuses[0].lower() == 'completed':
# #                             reason_text = 'AllLegsCompleted'
# #                             exit_time_to_use = None
# #                             if 'Exit Time' in group.columns:
# #                                 for exit_time, exit_type in zip(group['Exit Time'], group.get('Exit Type', [])):
# #                                     if pd.isna(exit_time):
# #                                         continue
# #                                     normalized_exit_time = str(exit_time).replace('.', ':').strip()
# #                                     matching_rows = df_grid[
# #                                         (df_grid['Option Portfolio'] == portfolio) &
# #                                         (df_grid['Timestamp'].astype(str).str.contains(normalized_exit_time))
# #                                     ]
# #                                     if not matching_rows.empty:
# #                                         reason_text += f", {exit_type.strip()}"
# #                                         exit_time_to_use = exit_time
# #                                         break
# #                             completed_list.append({
# #                                 'Option Portfolio': portfolio,
# #                                 'Reason': reason_text,
# #                                 'Time': exit_time_to_use
# #                             })
# #     if completed_list:
# #         completed_df = pd.DataFrame(completed_list)
# #         final_df = pd.concat([final_df, completed_df], ignore_index=True)
# #     logger.info("Cleaning Reason texts")
# #     def clean_reason(text):
# #         if pd.isna(text):
# #             return text
# #         text = str(text)
# #         match = re.search(r'(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)', text, re.IGNORECASE)
# #         if match:
# #             return match.group(1)
# #         if 'AllLegsCompleted' in text:
# #             text = text.replace('AllLegsCompleted,', '').strip()
# #             text = text.replace('AllLegsCompleted', '').strip()
# #         return text.strip()
# #     final_df['Reason'] = final_df['Reason'].apply(clean_reason)
# #     logger.info("Cleaning Time column")
# #     final_df['Time'] = final_df['Time'].astype(str).str.strip().replace('nan', None)
# #     logger.info("Portfolio data processing completed successfully")
# #     return final_df
# # # Function to process PNL data
# # def process_data(df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo):
# #     logger.info("Starting PNL data processing")
# #     try:
# #         logger.info("Positions DataFrame provided")
# #         required_columns = ['Exchange', 'Symbol', 'Net Qty', 'Buy Avg Price', 'Sell Avg Price', 'Sell Qty', 'Buy Qty', 'Realized Profit', 'Unrealized Profit']
# #         missing_columns = [col for col in required_columns if col not in df.columns]
# #         if missing_columns:
# #             raise ValueError(f"Missing columns in positions file: {missing_columns}")
# #         mask = df["Exchange"].isin(["NFO", "BFO"])
# #         if 'Symbol' in df.columns:
# #             df.loc[mask, "Symbol"] = df.loc[mask, "Symbol"].astype(str).str[-5:] + df.loc[mask, "Symbol"].astype(str).str[-8:-6]
# #         else:
# #             raise KeyError("Symbol column not found in positions file")
# #         df_nfo = df[df["Exchange"] == "NFO"].copy()
# #         df_bfo = df[df["Exchange"] == "BFO"].copy()
# #         logger.info("Data split into NFO and BFO")
# #         total_settlement_pnl_nfo = 0
# #         total_settlement_pnl_bfo = 0
# #         df_bhav_nfo = pd.DataFrame()
# #         df_bhav_bfo = pd.DataFrame()
# #         if include_settlement_nfo:
# #             if nfo_bhav_file is None:
# #                 raise ValueError("NFO bhavcopy file is required when settlement PNL for NFO is enabled")
# #             df_bhav_nfo = pd.read_csv(nfo_bhav_file)
# #             logger.info("NFO bhavcopy loaded successfully")
# #             if 'CONTRACT_D' not in df_bhav_nfo.columns:
# #                 raise ValueError("CONTRACT_D column not found in NFO bhavcopy")
# #             df_bhav_nfo["Date"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(\d{2}-[A-Z]{3}-\d{4})')
# #             df_bhav_nfo["Symbol"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'^(.*?)(\d{2}-[A-Z]{3}-\d{4})')[0]
# #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(PE\d+|CE\d+)$')
# #             df_bhav_nfo["Date"] = pd.to_datetime(df_bhav_nfo["Date"], format="%d-%b-%Y")
# #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["Strike_Type"].str.replace(
# #                 r'^(PE|CE)(\d+)$', r'\2\1', regex=True
# #             )
# #             target_symbol = "OPTIDXNIFTY"
# #             df_bhav_nfo = df_bhav_nfo[
# #                 (df_bhav_nfo["Date"] == pd.to_datetime(expiry_nfo)) &
# #                 (df_bhav_nfo["Symbol"] == target_symbol)
# #             ]
# #             df_nfo["Strike_Type"] = df_nfo["Symbol"].str.extract(r'(\d+[A-Z]{2})$')
# #             df_nfo['Strike'] = df_nfo['Strike_Type'].str[:-2].astype(float, errors='ignore')
# #             df_nfo['Option_Type'] = df_nfo['Strike_Type'].str[-2:]
# #             df_nfo = df_nfo.merge(
# #                 df_bhav_nfo[["Strike_Type", "SETTLEMENT"]],
# #                 on="Strike_Type",
# #                 how="left"
# #             )
# #             logger.info("NFO data merged with bhavcopy")
# #         if include_settlement_bfo:
# #             if bfo_bhav_file is None:
# #                 raise ValueError("BFO bhavcopy file is required when settlement PNL for BFO is enabled")
# #             df_bhav_bfo = pd.read_csv(bfo_bhav_file)
# #             logger.info("BFO bhavcopy loaded successfully")
# #             if 'Market Summary Date' not in df_bhav_bfo.columns or 'Expiry Date' not in df_bhav_bfo.columns or 'Series Code' not in df_bhav_bfo.columns:
# #                 raise ValueError("Required columns missing in BFO bhavcopy")
# #             df_bhav_bfo["Date"] = pd.to_datetime(df_bhav_bfo["Market Summary Date"], format="%d %b %Y", errors="coerce")
# #             df_bhav_bfo["Expiry Date"] = pd.to_datetime(df_bhav_bfo["Expiry Date"], format="%d %b %Y", errors="coerce")
# #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Series Code"].astype(str).str[-7:]
# #             df_bhav_bfo = df_bhav_bfo[
# #                 (df_bhav_bfo["Expiry Date"] == pd.to_datetime(expiry_bfo))
# #             ]
# #             df_bfo["Symbol"] = df_bfo["Symbol"].astype(str).str.strip()
# #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Symbols"].astype(str).str.strip()
# #             bhav_mapping = df_bhav_bfo.drop_duplicates(subset="Symbols", keep="last").set_index("Symbols")["Close Price"]
# #             df_bfo["Close Price"] = df_bfo["Symbol"].map(bhav_mapping)
# #             logger.info("BFO data processed")
# #         logger.info("Calculating Realized PNL for NFO")
# #         conditions = [
# #             df_nfo["Net Qty"] == 0,
# #             df_nfo["Net Qty"] > 0,
# #             df_nfo["Net Qty"] < 0
# #         ]
# #         choices = [
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Buy Qty"]
# #         ]
# #         df_nfo["Calculated_Realized_PNL"] = np.select(conditions, choices, default=0)
# #         df_nfo["Matching_Realized"] = df_nfo["Realized Profit"] == df_nfo["Calculated_Realized_PNL"]
# #         df_nfo["Matching_Realized"] = df_nfo["Matching_Realized"].replace({True: "TRUE", False: ""})
# #         if include_settlement_nfo:
# #             logger.info("Calculating Settlement PNL for NFO")
# #             df_nfo["Calculated_Settlement_PNL"] = np.select(
# #                 [
# #                     df_nfo["Net Qty"] > 0,
# #                     df_nfo["Net Qty"] < 0
# #                 ],
# #                 [
# #                     (df_nfo["SETTLEMENT"] - df_nfo["Buy Avg Price"]) * abs(df_nfo["Net Qty"]),
# #                     (df_nfo["Sell Avg Price"] - df_nfo["SETTLEMENT"]) * abs(df_nfo["Net Qty"])
# #                 ],
# #                 default=0
# #             )
# #             df_nfo["Matching_Settlement"] = df_nfo["Unrealized Profit"] == df_nfo["Calculated_Settlement_PNL"]
# #             df_nfo["Matching_Settlement"] = df_nfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# #             total_settlement_pnl_nfo = df_nfo["Calculated_Settlement_PNL"].fillna(0).sum()
# #         else:
# #             df_nfo["Calculated_Settlement_PNL"] = 0
# #             df_nfo["Matching_Settlement"] = ""
# #         total_realized_pnl_nfo = df_nfo["Calculated_Realized_PNL"].fillna(0).sum()
# #         logger.info("Calculating Realized PNL for BFO")
# #         conditions_bfo = [
# #             df_bfo["Net Qty"] == 0,
# #             df_bfo["Net Qty"] > 0,
# #             df_bfo["Net Qty"] < 0
# #         ]
# #         choices_bfo = [
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Buy Qty"]
# #         ]
# #         df_bfo["Calculated_Realized_PNL"] = np.select(conditions_bfo, choices_bfo, default=0)
# #         df_bfo["Matching_Realized"] = df_bfo["Realized Profit"] == df_bfo["Calculated_Realized_PNL"]
# #         df_bfo["Matching_Realized"] = df_bfo["Matching_Realized"].replace({True: "TRUE", False: ""})
# #         if include_settlement_bfo:
# #             logger.info("Calculating Settlement PNL for BFO")
# #             long_condition = df_bfo["Net Qty"] > 0
# #             df_bfo.loc[long_condition, "Calculated_Settlement_PNL"] = (
# #                 (df_bfo["Close Price"] - df_bfo["Buy Avg Price"]) * abs(df_bfo["Net Qty"])
# #             )
# #             short_condition = df_bfo["Net Qty"] < 0
# #             df_bfo.loc[short_condition, "Calculated_Settlement_PNL"] = (
# #                 (df_bfo["Sell Avg Price"] - df_bfo["Close Price"]) * abs(df_bfo["Net Qty"])
# #             )
# #             df_bfo.loc[df_bfo["Net Qty"] == 0, "Calculated_Settlement_PNL"] = 0
# #             df_bfo["Matching_Settlement"] = df_bfo["Unrealized Profit"] == df_bfo["Calculated_Settlement_PNL"]
# #             df_bfo["Matching_Settlement"] = df_bfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# #             total_settlement_pnl_bfo = df_bfo["Calculated_Settlement_PNL"].fillna(0).sum()
# #         else:
# #             df_bfo["Calculated_Settlement_PNL"] = 0
# #             df_bfo["Matching_Settlement"] = ""
# #         total_realized_pnl_bfo = df_bfo["Calculated_Realized_PNL"].fillna(0).sum()
# #         overall_realized = total_realized_pnl_nfo + total_realized_pnl_bfo
# #         overall_settlement = total_settlement_pnl_nfo + total_settlement_pnl_bfo
# #         grand_total = overall_realized + overall_settlement
# #         logger.info("PNL data processing completed successfully")
# #         return {
# #             "df_nfo": df_nfo,
# #             "df_bfo": df_bfo,
# #             "total_realized_nfo": total_realized_pnl_nfo,
# #             "total_settlement_nfo": total_settlement_pnl_nfo,
# #             "total_realized_bfo": total_realized_pnl_bfo,
# #             "total_settlement_bfo": total_settlement_pnl_bfo,
# #             "overall_realized": overall_realized,
# #             "overall_settlement": overall_settlement,
# #             "grand_total": grand_total
# #         }
# #     except Exception as e:
# #         logger.error(f"Error in process_data: {str(e)}")
# #         raise
# # # Function to get Excel download link
# # def get_excel_download_link(df, filename):
# #     output = BytesIO()
# #     with pd.ExcelWriter(output, engine='openpyxl') as writer:
# #         df.to_excel(writer, index=False, sheet_name='PNL Data')
# #         worksheet = writer.sheets['PNL Data']
# #         for row in worksheet.rows:
# #             for cell in row:
# #                 cell.alignment = Alignment(horizontal='center')
# #         for cell in worksheet[1]:
# #             cell.font = Font(bold=True, color="FFFFFF")
# #             cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
# #         for row in worksheet.rows:
# #             for cell in row:
# #                 cell.border = Border(left=Side(style='thin'),
# #                                      right=Side(style='thin'),
# #                                      top=Side(style='thin'),
# #                                      bottom=Side(style='thin'))
# #     excel_data = output.getvalue()
# #     b64 = base64.b64encode(excel_data).decode()
# #     return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}.xlsx</a>'
# # # Function to get CSV download link
# # def get_csv_download_link(df, filename):
# #     output = BytesIO()
# #     df.to_csv(output, index=False)
# #     csv_data = output.getvalue()
# #     b64 = base64.b64encode(csv_data).decode()
# #     return f'<a href="data:text/csv;base64,{b64}" download="{filename}.csv" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}.csv</a>'
# # # Run the app
# # if __name__ == "__main__":
# #     st.write(f"DEBUG: Starting Streamlit app at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
# #     run()



# import streamlit as st
# import pandas as pd
# import numpy as np
# import re
# import base64
# from io import BytesIO
# import logging
# import openpyxl
# from openpyxl.styles import Alignment, Border, Side, Font
# from datetime import datetime

# # Set up logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Check Streamlit version for compatibility
# try:
#     import streamlit
#     logger.info(f"Streamlit version: {streamlit.__version__}")
# except ImportError:
#     st.error("Streamlit is not installed. Please install it using `pip install streamlit`.")
#     st.stop()

# # ===================== CUSTOM CSS & STYLING =====================
# st.markdown("""
# <link href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css" rel="stylesheet">
# <style>
# body {
#     font-family: 'Inter', sans-serif;
#     background: linear-gradient(135deg, #f3f4f6, #e5e7eb);
# }
# .stButton>button {
#     background: linear-gradient(45deg, #3b82f6, #60a5fa);
#     color: white;
#     border: none;
#     padding: 0.75rem 1.5rem;
#     border-radius: 0.5rem;
#     font-weight: 600;
#     transition: all 0.3s ease;
#     width: 100%;
# }
# .stButton>button:hover {
#     background: linear-gradient(45deg, #2563eb, #3b82f6);
#     transform: translateY(-2px);
#     box-shadow: 0 4px 6px rgba(0,0,0,0.2);
# }
# .stDateInput input {
#     border: 2px solid #3b82f6;
#     border-radius: 0.5rem;
#     padding: 0.5rem;
#     background: #ffffff;
#     color: #1f2937;
#     font-size: 1rem;
#     transition: all 0.3s ease;
# }
# .stDateInput input:focus {
#     outline: none;
#     border-color: #2563eb;
#     box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
#     background: #f8fafc;
# }
# .stFileUploader button {
#     background: linear-gradient(45deg, #10b981, #34d399);
#     color: white;
#     border-radius: 0.5rem;
#     padding: 0.75rem;
# }
# .stFileUploader button:hover {
#     background: linear-gradient(45deg, #059669, #10b981);
# }
# .stCheckbox label {
#     font-size: 1rem;
#     color: #1f2937;
# }
# .metric-card {
#     background: #ffffff;
#     padding: 1.5rem;
#     border-radius: 0.75rem;
#     box-shadow: 0 4px 6px rgba(0,0,0,0.1);
#     text-align: center;
#     transition: transform 0.3s ease;
# }
# .metric-card:hover {
#     transform: translateY(-5px);
#     box-shadow: 0 6px 12px rgba(0,0,0,0.15);
# }
# .metric-label {
#     font-size: 1.1rem;
#     color: #6b7280;
#     margin-bottom: 0.5rem;
# }
# .metric-value {
#     font-size: 1.75rem;
#     font-weight: 700;
# }
# .stTabs [data-baseweb="tab"] {
#     font-size: 1.1rem;
#     font-weight: 600;
#     padding: 0.75rem 1.5rem;
#     border-radius: 0.5rem;
#     transition: all 0.3s ease;
# }
# .stTabs [data-baseweb="tab"]:hover {
#     background: #e5e7eb;
# }
# .insights-box {
#     background: #ffffff;
#     padding: 1.5rem;
#     border-radius: 0.5rem;
#     box-shadow: 0 4px 6px rgba(0,0,0,0.1);
#     margin-top: 1.5rem;
# }
# .chart-container {
#     background: #ffffff;
#     padding: 1.5rem;
#     border-radius: 0.75rem;
#     box-shadow: 0 4px 6px rgba(0,0,0,0.1);
#     margin-bottom: 1.5rem;
# }
# .header-text {
#     font-size: 2.25rem;
#     font-weight: 800;
#     color: #1f2937;
#     text-align: center;
#     margin-bottom: 1rem;
# }
# .subheader-text {
#     font-size: 1.25rem;
#     color: #4b5563;
#     text-align: center;
#     margin-bottom: 2rem;
# }
# footer { visibility: hidden; }
# </style>
# """, unsafe_allow_html=True)

# # ===================== ALL FUNCTIONS FIRST (FIXED ORDER) =====================

# def process_portfolio_data(gridlog_file, summary_file):
#     """
#     Exact same logic as your standalone script.
#     Takes uploaded GridLog (CSV/XLSX) and Summary (XLSX) files.
#     Returns final_df and dynamic filename.
#     """
#     # --- Step 1: Load GridLog file ---
#     if gridlog_file.name.endswith('.csv'):
#         df_grid = pd.read_csv(gridlog_file)
#     elif gridlog_file.name.endswith('.xlsx'):
#         df_grid = pd.read_excel(gridlog_file)
#     else:
#         raise ValueError("Unsupported GridLog file type. Use CSV or Excel.")

#     # --- Step 2: Clean column names ---
#     df_grid.columns = df_grid.columns.str.strip()

#     # --- Step 3: Filter for Combined SL / Trail Target messages ---
#     mask = df_grid['Message'].str.contains(r'Combined SL:|Combined trail target:', case=False, na=False)
#     filtered_grid = df_grid.loc[mask, ['Message', 'Option Portfolio', 'Timestamp']].dropna(subset=['Option Portfolio'])

#     # Keep only Option Portfolios that appear more than once for same message type
#     filtered_grid['MessageType'] = filtered_grid['Message'].str.extract(r'(Combined SL|Combined trail target)', flags=re.IGNORECASE)
#     duplicate_mask = filtered_grid.duplicated(subset=['Option Portfolio', 'MessageType'], keep=False)
#     filtered_grid = filtered_grid[duplicate_mask]

#     # Group GridLog results
#     summary_grid = (
#         filtered_grid.groupby('Option Portfolio').agg({
#             'Message': lambda x: ', '.join(x.unique()),
#             'Timestamp': 'max'
#         }).reset_index()
#         .rename(columns={'Message': 'Reason', 'Timestamp': 'Time'})
#     )

#     # --- Step 4: Process Summary Excel File ---
#     xl = pd.ExcelFile(summary_file)
#     summary_list = []

#     for sheet_name in xl.sheet_names:
#         if "legs" in sheet_name.lower():
#             df_leg = xl.parse(sheet_name)
#             df_leg.columns = df_leg.columns.str.strip()

#             if {'Exit Type', 'Portfolio Name', 'Exit Time'}.issubset(df_leg.columns):
#                 onsqoff_df = df_leg[df_leg['Exit Type'].astype(str).str.strip() == 'OnSqOffTime']

#                 if not onsqoff_df.empty:
#                     grouped = onsqoff_df.groupby('Portfolio Name')['Exit Time'].max().reset_index()
#                     for _, row in grouped.iterrows():
#                         summary_list.append({
#                             'Option Portfolio': row['Portfolio Name'],
#                             'Reason': 'OnSqOffTime',
#                             'Time': row['Exit Time']
#                         })

#     summary_summary = pd.DataFrame(summary_list)

#     # --- Step 5: Combine GridLog and Summary results ---
#     final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)

#     final_df = (
#         final_df.groupby('Option Portfolio').agg({
#             'Reason': lambda x: ', '.join(sorted(set(x))),
#             'Time': 'last'
#         }).reset_index()
#     )

#     # --- Step 6: Add completed portfolios (AllLegsCompleted) ---
#     completed_list = []
#     grid_portfolios = df_grid['Option Portfolio'].dropna().unique()

#     for sheet_name in xl.sheet_names:
#         if "legs" in sheet_name.lower():
#             df_leg = xl.parse(sheet_name)
#             df_leg.columns = df_leg.columns.str.strip()

#             if 'Portfolio Name' in df_leg.columns and 'Status' in df_leg.columns:
#                 for portfolio, group in df_leg.groupby('Portfolio Name'):
#                     if (portfolio not in final_df['Option Portfolio'].values 
#                         and portfolio in grid_portfolios):
                        
#                         statuses = group['Status'].astype(str).str.strip().unique()
#                         if len(statuses) == 1 and statuses[0].lower() == 'completed':
#                             reason_text = 'AllLegsCompleted'
#                             exit_time_to_use = None

#                             if 'Exit Time' in group.columns:
#                                 for exit_time, exit_type in zip(group['Exit Time'], group.get('Exit Type', [])):
#                                     if pd.isna(exit_time):
#                                         continue
#                                     normalized_exit_time = str(exit_time).replace('.', ':').strip()
#                                     matching_rows = df_grid[
#                                         (df_grid['Option Portfolio'] == portfolio) &
#                                         (df_grid['Timestamp'].astype(str).str.contains(normalized_exit_time))
#                                     ]
#                                     if not matching_rows.empty:
#                                         reason_text += f", {exit_type.strip()}"
#                                         exit_time_to_use = exit_time
#                                         break

#                             completed_list.append({
#                                 'Option Portfolio': portfolio,
#                                 'Reason': reason_text,
#                                 'Time': exit_time_to_use
#                             })

#     if completed_list:
#         completed_df = pd.DataFrame(completed_list)
#         final_df = pd.concat([final_df, completed_df], ignore_index=True)

#     # --- Step 7: Clean Reason Texts ---
#     def clean_reason(text):
#         if pd.isna(text):
#             return text
#         text = str(text)

#         # Case 1: Keep only main Combined SL / Trail Target reason
#         match = re.search(r'(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)', text, re.IGNORECASE)
#         if match:
#             return match.group(1)

#         # Case 2: Remove "AllLegsCompleted" prefix if present
#         if 'AllLegsCompleted' in text:
#             text = text.replace('AllLegsCompleted,', '').replace('AllLegsCompleted', '').strip()

#         return text.strip()

#     final_df['Reason'] = final_df['Reason'].apply(clean_reason)

#     # --- Step 8: Build dynamic output file name ---
#     filename = gridlog_file.name
#     match = re.search(r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', filename)
#     if match:
#         raw_date = match.group(1)
#         parts = raw_date.split()
#         formatted_date = f"{parts[0]} {parts[1].lower()}"
#     else:
#         formatted_date = "unknown_date"

#     output_filename = f"completed portfolio of {formatted_date}.csv"

#     # --- Step 9: Final cleanup ---
#     final_df['Time'] = final_df['Time'].astype(str).str.strip().replace('nan', None)

#     return final_df, output_filename

# def process_data(df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo,
#                  include_settlement_nfo, include_settlement_bfo):
#     # Your original PNL logic – unchanged
#     logger.info("Starting PNL data processing")
#     try:
#         required_columns = ['Exchange', 'Symbol', 'Net Qty', 'Buy Avg Price', 'Sell Avg Price',
#                             'Sell Qty', 'Buy Qty', 'Realized Profit', 'Unrealized Profit']
#         missing = [c for c in required_columns if c not in df.columns]
#         if missing:
#             raise ValueError(f"Missing columns: {missing}")

#         mask = df["Exchange"].isin(["NFO", "BFO"])
#         df.loc[mask, "Symbol"] = df.loc[mask, "Symbol"].astype(str).str[-5:] + df.loc[mask, "Symbol"].astype(str).str[-8:-6]

#         df_nfo = df[df["Exchange"] == "NFO"].copy()
#         df_bfo = df[df["Exchange"] == "BFO"].copy()

#         total_realized_nfo = total_realized_bfo = 0
#         total_settlement_nfo = total_settlement_bfo = 0

#         # NFO Realized
#         cond_nfo = [df_nfo["Net Qty"] == 0, df_nfo["Net Qty"] > 0, df_nfo["Net Qty"] < 0]
#         choice_nfo = [
#             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
#             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
#             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Buy Qty"]
#         ]
#         df_nfo["Calculated_Realized_PNL"] = np.select(cond_nfo, choice_nfo, default=0)
#         total_realized_nfo = df_nfo["Calculated_Realized_PNL"].fillna(0).sum()

#         # BFO Realized
#         cond_bfo = [df_bfo["Net Qty"] == 0, df_bfo["Net Qty"] > 0, df_bfo["Net Qty"] < 0]
#         choice_bfo = [
#             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
#             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
#             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Buy Qty"]
#         ]
#         df_bfo["Calculated_Realized_PNL"] = np.select(cond_bfo, choice_bfo, default=0)
#         total_realized_bfo = df_bfo["Calculated_Realized_PNL"].fillna(0).sum()

#         # Settlement logic (kept exactly as you had)
#         if include_settlement_nfo and nfo_bhav_file:
#             df_bhav_nfo = pd.read_csv(nfo_bhav_file)
#             df_bhav_nfo["Date"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(\d{2}-[A-Z]{3}-\d{4})')
#             df_bhav_nfo["Symbol"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'^(.*?)(\d{2}-[A-Z]{3}-\d{4})')[0]
#             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(PE\d+|CE\d+)$')
#             df_bhav_nfo["Date"] = pd.to_datetime(df_bhav_nfo["Date"], format="%d-%b-%Y")
#             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["Strike_Type"].str.replace(r'^(PE|CE)(\d+)$', r'\2\1', regex=True)
#             df_bhav_nfo = df_bhav_nfo[(df_bhav_nfo["Date"] == pd.to_datetime(expiry_nfo)) & (df_bhav_nfo["Symbol"] == "OPTIDXNIFTY")]
#             df_nfo["Strike_Type"] = df_nfo["Symbol"].str.extract(r'(\d+[A-Z]{2})$')
#             df_nfo = df_nfo.merge(df_bhav_nfo[["Strike_Type", "SETTLEMENT"]], on="Strike_Type", how="left")
#             df_nfo["Calculated_Settlement_PNL"] = np.select(
#                 [df_nfo["Net Qty"] > 0, df_nfo["Net Qty"] < 0],
#                 [(df_nfo["SETTLEMENT"] - df_nfo["Buy Avg Price"]) * abs(df_nfo["Net Qty"]),
#                  (df_nfo["Sell Avg Price"] - df_nfo["SETTLEMENT"]) * abs(df_nfo["Net Qty"])],
#                 default=0)
#             total_settlement_nfo = df_nfo["Calculated_Settlement_PNL"].fillna(0).sum()

#         if include_settlement_bfo and bfo_bhav_file:
#             df_bhav_bfo = pd.read_csv(bfo_bhav_file)
#             df_bhav_bfo["Expiry Date"] = pd.to_datetime(df_bhav_bfo["Expiry Date"], format="%d %b %Y", errors="coerce")
#             df_bhav_bfo = df_bhav_bfo[df_bhav_bfo["Expiry Date"] == pd.to_datetime(expiry_bfo)]
#             df_bhav_bfo["Symbols"] = df_bhav_bfo["Series Code"].astype(str).str[-7:]
#             mapping = df_bhav_bfo.drop_duplicates("Symbols").set_index("Symbols")["Close Price"]
#             df_bfo["Close Price"] = df_bfo["Symbol"].astype(str).str.strip().map(mapping)
#             df_bfo["Calculated_Settlement_PNL"] = 0
#             df_bfo.loc[df_bfo["Net Qty"] > 0, "Calculated_Settlement_PNL"] = (df_bfo["Close Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Net Qty"].abs()
#             df_bfo.loc[df_bfo["Net Qty"] < 0, "Calculated_Settlement_PNL"] = (df_bfo["Sell Avg Price"] - df_bfo["Close Price"]) * df_bfo["Net Qty"].abs()
#             total_settlement_bfo = df_bfo["Calculated_Settlement_PNL"].fillna(0).sum()

#         overall_realized = total_realized_nfo + total_realized_bfo
#         overall_settlement = total_settlement_nfo + total_settlement_bfo
#         grand_total = overall_realized + overall_settlement

#         return {
#             "total_realized_nfo": total_realized_nfo,
#             "total_settlement_nfo": total_settlement_nfo,
#             "total_realized_bfo": total_realized_bfo,
#             "total_settlement_bfo": total_settlement_bfo,
#             "overall_realized": overall_realized,
#             "overall_settlement": overall_settlement,
#             "grand_total": grand_total
#         }
#     except Exception as e:
#         logger.error(f"Error in process_data: {e}")
#         raise


# def get_excel_download_link(df, filename):
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='PNL Data')
#         ws = writer.sheets['PNL Data']
#         for row in ws.rows:
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center')
#                 cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                                      top=Side(style='thin'), bottom=Side(style='thin'))
#         for cell in ws[1]:
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", fill_type="solid")
#     b64 = base64.b64encode(output.getvalue()).decode()
#     return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Download {filename}.xlsx</a>'


# def get_csv_download_link(df, filename):
#     csv = df.to_csv(index=False).encode()
#     b64 = base64.b64encode(csv).decode()
#     return f'<a href="data:text/csv;base64,{b64}" download="{filename}">Download {filename}</a>'

# # ===================== TABS =====================
# tab1, tab2 = st.tabs(["Full PNL Calculation", "Portfolio Exit Analysis"])

# # ===================== TAB 1: PNL CALCULATION =====================
# with tab1:
#     st.markdown('<h1 class="header-text">Ultimate Financial PNL Dashboard</h1>', unsafe_allow_html=True)
#     st.markdown('<p class="subheader-text">Analyze your portfolio with real-time insights, interactive charts, and comprehensive data exports for smarter trading decisions.</p>', unsafe_allow_html=True)
    
#     # Input Section for PNL Dashboard
#     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Upload Data</h2>', unsafe_allow_html=True)
#     with st.container():
#         positions_file = st.file_uploader("Positions CSV", type="csv", help="Upload VS20 22 AUG 2025 POSITIONS(EOD).csv", key="positions_upload")
#         selected_user = None
#         if positions_file:
#             if 'positions_df' not in st.session_state or st.session_state.get('positions_file_name') != positions_file.name:
#                 st.session_state.positions_df = pd.read_csv(positions_file)
#                 st.session_state.positions_file_name = positions_file.name
#             df = st.session_state.positions_df
#             if 'UserID' in df.columns:
#                 users = sorted(df['UserID'].unique().tolist())
#                 selected_user = st.selectbox("Select User", users, key="selected_user")
#             else:
#                 st.error("'UserID' column not found in positions file.")
        
#         # Create two columns for the checkboxes (same row)
#         checkbox_col1, checkbox_col2 = st.columns(2)
#         with checkbox_col1:
#             include_settlement_nfo = st.checkbox(
#                 "Include Settlement PNL for NFO",
#                 value=True,
#                 help="Uncheck to exclude settlement PNL for NFO",
#                 key="nfo_settlement"
#             )
#         with checkbox_col2:
#             include_settlement_bfo = st.checkbox(
#                 "Include Settlement PNL for BFO",
#                 value=True,
#                 help="Uncheck to exclude settlement PNL for BFO",
#                 key="bfo_settlement"
#             )
        
#         # Layout for file uploaders
#         col1, col2 = st.columns(2)
#         with col1:
#             if include_settlement_nfo:
#                 nfo_bhav_file = st.file_uploader(
#                     "NFO Bhavcopy",
#                     type="csv",
#                     help="Upload op220825.csv",
#                     key="nfo_upload"
#                 )
#             else:
#                 nfo_bhav_file = None
#                 st.info("NFO Bhavcopy not required when settlement PNL for NFO is disabled.")
#         with col2:
#             if include_settlement_bfo:
#                 bfo_bhav_file = st.file_uploader(
#                     "BFO Bhavcopy",
#                     type="csv",
#                     help="Upload MS_20250822-01.csv",
#                     key="bfo_upload"
#                 )
#             else:
#                 bfo_bhav_file = None
#                 st.info("BFO Bhavcopy not required when settlement PNL for BFO is disabled.")
        
#         st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Expiry Dates</h2>', unsafe_allow_html=True)
#         with st.container():
#             col3, col4 = st.columns(2)
#             with col3:
#                 expiry_nfo = st.date_input("NFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="nfo_expiry", disabled=not include_settlement_nfo)
#             with col4:
#                 expiry_bfo = st.date_input("BFO Expiry Date", value=datetime.now().date(), help="Format: YYYY-MM-DD", key="bfo_expiry", disabled=not include_settlement_bfo)
        
#         process_button = st.button("Process Data", key="process_button")
#         if process_button:
#             if positions_file and selected_user:
#                 if (include_settlement_nfo and nfo_bhav_file is None) or (include_settlement_bfo and bfo_bhav_file is None):
#                     st.error("Please upload all required CSV files for enabled settlement calculations.")
#                 else:
#                     try:
#                         with st.spinner("Processing your data..."):
#                             filtered_df = st.session_state.positions_df[st.session_state.positions_df['UserID'] == selected_user]
#                             results = process_data(filtered_df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo)
                        
#                         # Key Metrics Section
#                         st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">Key Financial Metrics</h2>', unsafe_allow_html=True)
#                         col1, col2, col3 = st.columns(3, gap="medium")
#                         with col1:
#                             color = '#34D399' if results["overall_realized"] > 0 else '#F87171' if results["overall_realized"] < 0 else '#9CA3AF'
#                             st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_realized"]:,.2f}</p></div>', unsafe_allow_html=True)
#                         if include_settlement_nfo or include_settlement_bfo:
#                             with col2:
#                                 color = '#34D399' if results["overall_settlement"] > 0 else '#F87171' if results["overall_settlement"] < 0 else '#9CA3AF'
#                                 st.markdown(f'<div class="metric-card"><p class="metric-label">Overall Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["overall_settlement"]:,.2f}</p></div>', unsafe_allow_html=True)
#                             with col3:
#                                 color = '#34D399' if results["grand_total"] > 0 else '#F87171' if results["grand_total"] < 0 else '#9CA3AF'
#                                 st.markdown(f'<div class="metric-card"><p class="metric-label">Grand Total PNL</p><p class="metric-value" style="color:{color};">₹{results["grand_total"]:,.2f}</p></div>', unsafe_allow_html=True)
#                         col4, col5, col6, col7 = st.columns(4, gap="medium")
#                         with col4:
#                             color = '#34D399' if results["total_realized_nfo"] > 0 else '#F87171' if results["total_realized_nfo"] < 0 else '#9CA3AF'
#                             st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
#                         if include_settlement_nfo:
#                             with col5:
#                                 color = '#34D399' if results["total_settlement_nfo"] > 0 else '#F87171' if results["total_settlement_nfo"] < 0 else '#9CA3AF'
#                                 st.markdown(f'<div class="metric-card"><p class="metric-label">NFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_nfo"]:,.2f}</p></div>', unsafe_allow_html=True)
#                         with col6:
#                             color = '#34D399' if results["total_realized_bfo"] > 0 else '#F87171' if results["total_realized_bfo"] < 0 else '#9CA3AF'
#                             st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Realized PNL</p><p class="metric-value" style="color:{color};">₹{results["total_realized_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)
#                         if include_settlement_bfo:
#                             with col7:
#                                 color = '#34D399' if results["total_settlement_bfo"] > 0 else '#F87171' if results["total_settlement_bfo"] < 0 else '#9CA3AF'
#                                 st.markdown(f'<div class="metric-card"><p class="metric-label">BFO Settlement PNL</p><p class="metric-value" style="color:{color};">₹{results["total_settlement_bfo"]:,.2f}</p></div>', unsafe_allow_html=True)
                        
#                         # Download Results
#                         st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">Download Results</h2>', unsafe_allow_html=True)
#                         col_download1 = st.columns(1)[0]
#                         with col_download1:
#                             pnl_breakdown_df = pd.DataFrame({
#                                 'Date': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
#                                 'NFO Realized PNL': [results["total_realized_nfo"]],
#                                 'NFO Settlement PNL': [results["total_settlement_nfo"] if include_settlement_nfo else 0],
#                                 'Nfo total': [results["total_realized_nfo"] + (results["total_settlement_nfo"] if include_settlement_nfo else 0)],
#                                 'BFO Realized PNL': [results["total_realized_bfo"]],
#                                 'BFO Settlement PNL': [results["total_settlement_bfo"] if include_settlement_bfo else 0],
#                                 'BFO total': [results["total_realized_bfo"] + (results["total_settlement_bfo"] if include_settlement_bfo else 0)],
#                                 'Grand Total PNL': [results["grand_total"]]
#                             })
#                             st.markdown(get_excel_download_link(pnl_breakdown_df, "A19_data"), unsafe_allow_html=True)
#                         st.success("PNL data processed successfully!")
#                     except Exception as e:
#                         logger.error(f"Error in process_data: {str(e)}")
#                         st.error(f"Error processing data: {str(e)}")
#                         st.info("Please ensure all uploaded CSV files have the correct columns and data format. Check the log for details.")
#             else:
#                 st.error("Please upload the positions CSV file and select a user to proceed.")

# # ===================== TAB 2: PORTFOLIO ANALYSIS (UPDATED LOGIC) =====================
# with tab2:
#     st.markdown('<hr class="my-8 border-gray-300">', unsafe_allow_html=True)
#     st.markdown('<h1 class="header-text">Portfolio Analysis</h1>', unsafe_allow_html=True)
#     st.markdown('<p class="subheader-text">Upload GridLog and Summary files to analyze portfolio exit reasons and timestamps.</p>', unsafe_allow_html=True)
#     st.info("Upload the required files below and click 'Process Portfolio Data' to view results.")
#     st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Upload Portfolio Data</h2>', unsafe_allow_html=True)
    
#     with st.container():
#         col_grid, col_summary = st.columns(2)
#         with col_grid:
#             gridlog_file = st.file_uploader(
#                 "GridLog File",
#                 type=["csv", "xlsx"],
#                 help="Upload VS20 18 NOV 2025 GridLog.csv or .xlsx",
#                 key="gridlog_upload"
#             )
#         with col_summary:
#             summary_file = st.file_uploader(
#                 "Summary Excel File",
#                 type="xlsx",
#                 help="Upload VS20 18 NOV 2025 SUMMARY.xlsx",
#                 key="summary_upload"
#             )
    
#     process_portfolio_button = st.button("Process Portfolio Data", key="process_portfolio_button")
    
#     if process_portfolio_button:
#         if gridlog_file and summary_file:
#             try:
#                 with st.spinner("Processing portfolio data..."):
#                     final_df, output_filename = process_portfolio_data(gridlog_file, summary_file)
                
#                 st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">Portfolio Analysis Results</h2>', unsafe_allow_html=True)
#                 st.markdown('<div class="insights-box">', unsafe_allow_html=True)
#                 st.write(final_df)
#                 st.markdown('</div>', unsafe_allow_html=True)
#                 st.success("Portfolio data processed successfully!")
                
#                 # Dynamic Download Link
#                 st.markdown('<h2 class="text-xl font-bold text-gray-800 dark:text-gray-200 mt-8 mb-4">Download Portfolio Results</h2>', unsafe_allow_html=True)
#                 st.markdown(get_csv_download_link(final_df, output_filename), unsafe_allow_html=True)
                
#             except Exception as e:
#                 logger.error(f"Error in process_portfolio_data: {str(e)}")
#                 st.error(f"Error processing portfolio data: {str(e)}")
#                 st.info("Please ensure the uploaded files have the correct format and columns.")
#         else:
#             st.error("Please upload both GridLog and Summary files to proceed.")

# # # ===================== UPDATED process_portfolio_data FUNCTION (EXACT NEW LOGIC) =====================
# # def process_portfolio_data(gridlog_file, summary_file):
# #     # Step 1: Load GridLog
# #     if gridlog_file.name.endswith('.csv'):
# #         df_grid = pd.read_csv(gridlog_file)
# #     elif gridlog_file.name.endswith('.xlsx'):
# #         df_grid = pd.read_excel(gridlog_file)
# #     else:
# #         raise ValueError("GridLog must be .csv or .xlsx")

# #     df_grid.columns = df_grid.columns.str.strip()

# #     # Step 3: Filter Combined SL / Trail Target + Keep only duplicates
# #     mask = df_grid['Message'].str.contains(r'Combined SL:|Combined trail target:', case=False, na=False)
# #     filtered_grid = df_grid.loc[mask, ['Message', 'Option Portfolio', 'Timestamp']].dropna(subset=['Option Portfolio'])

# #     filtered_grid['MessageType'] = filtered_grid['Message'].str.extract(r'(Combined SL|Combined trail target)', flags=re.IGNORECASE)
# #     duplicate_mask = filtered_grid.duplicated(subset=['Option Portfolio', 'MessageType'], keep=False)
# #     filtered_grid = filtered_grid[duplicate_mask]

# #     summary_grid = (
# #         filtered_grid.groupby('Option Portfolio')
# #         .agg({'Message': lambda x: ', '.join(x.unique()), 'Timestamp': 'max'})
# #         .reset_index()
# #         .rename(columns={'Message': 'Reason', 'Timestamp': 'Time'})
# #     )

# #     # Step 4: OnSqOffTime from Summary
# #     xl = pd.ExcelFile(summary_file)
# #     summary_list = []
# #     for sheet in xl.sheet_names:
# #         if "legs" in sheet.lower():
# #             df_leg = xl.parse(sheet)
# #             df_leg.columns = df_leg.columns.str.strip()
# #             if {'Exit Type', 'Portfolio Name', 'Exit Time'}.issubset(df_leg.columns):
# #                 onsqoff = df_leg[df_leg['Exit Type'].astype(str).str.strip() == 'OnSqOffTime']
# #                 if not onsqoff.empty:
# #                     grouped = onsqoff.groupby('Portfolio Name')['Exit Time'].max().reset_index()
# #                     for _, row in grouped.iterrows():
# #                         summary_list.append({
# #                             'Option Portfolio': row['Portfolio Name'],
# #                             'Reason': 'OnSqOffTime',
# #                             'Time': row['Exit Time']
# #                         })
# #     summary_summary = pd.DataFrame(summary_list)

# #     # Step 5: Combine
# #     final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)
# #     final_df = final_df.groupby('Option Portfolio').agg({
# #         'Reason': lambda x: ', '.join(sorted(set(x))),
# #         'Time': 'last'
# #     }).reset_index()

# #     # Step 6: AllLegsCompleted
# #     completed_list = []
# #     grid_portfolios = df_grid['Option Portfolio'].dropna().unique()

# #     for sheet in xl.sheet_names:
# #         if "legs" in sheet.lower():
# #             df_leg = xl.parse(sheet)
# #             df_leg.columns = df_leg.columns.str.strip()
# #             if {'Portfolio Name', 'Status'}.issubset(df_leg.columns):
# #                 for portfolio, group in df_leg.groupby('Portfolio Name'):
# #                     if (portfolio not in final_df['Option Portfolio'].values 
# #                         and portfolio in grid_portfolios
# #                         and group['Status'].astype(str).str.strip().eq('completed').all()):
                        
# #                         reason_text = 'AllLegsCompleted'
# #                         exit_time_to_use = None

# #                         if 'Exit Time' in group.columns:
# #                             for exit_time, exit_type in zip(group['Exit Time'], group.get('Exit Type', [])):
# #                                 if pd.isna(exit_time):
# #                                     continue
# #                                 norm_time = str(exit_time).replace('.', ':').strip()
# #                                 if df_grid[(df_grid['Option Portfolio'] == portfolio) & 
# #                                           (df_grid['Timestamp'].astype(str).str.contains(norm_time))].shape[0] > 0:
# #                                     reason_text += f", {exit_type.strip()}" if pd.notna(exit_type) else ""
# #                                     exit_time_to_use = exit_time
# #                                     break

# #                         completed_list.append({
# #                             'Option Portfolio': portfolio,
# #                             'Reason': reason_text,
# #                             'Time': exit_time_to_use
# #                         })

# #     if completed_list:
# #         final_df = pd.concat([final_df, pd.DataFrame(completed_list)], ignore_index=True)

# #     # Step 7: Clean Reason
# #     def clean_reason(text):
# #         if pd.isna(text):
# #             return text
# #         text = str(text)
# #         match = re.search(r'(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)', text, re.IGNORECASE)
# #         if match:
# #             return match.group(1)
# #         text = text.replace('AllLegsCompleted,', '').replace('AllLegsCompleted', '').strip()
# #         return text.strip()

# #     final_df['Reason'] = final_df['Reason'].apply(clean_reason)

# #     # Step 8: Dynamic filename
# #     match = re.search(r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', gridlog_file.name)
# #     date_str = match.group(1).split()
# #     formatted_date = f"{date_str[0]} {date_str[1].lower()}" if match else "unknown_date"
# #     output_filename = f"completed portfolio of {formatted_date}.csv"

# #     # Step 9: Final cleanup
# #     final_df['Time'] = final_df['Time'].astype(str).str.strip().replace('nan', None)

# #     return final_df, output_filename

# # # ===================== process_data (unchanged) =====================
# # def process_data(df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo, include_settlement_nfo, include_settlement_bfo):
# #     # ... [Your existing PNL logic remains 100% unchanged] ...
# #     # (kept as-is from your original code)
# #     logger.info("Starting PNL data processing")
# #     try:
# #         required_columns = ['Exchange', 'Symbol', 'Net Qty', 'Buy Avg Price', 'Sell Avg Price', 'Sell Qty', 'Buy Qty', 'Realized Profit', 'Unrealized Profit']
# #         missing_columns = [col for col in required_columns if col not in df.columns]
# #         if missing_columns:
# #             raise ValueError(f"Missing columns in positions file: {missing_columns}")
# #         mask = df["Exchange"].isin(["NFO", "BFO"])
# #         if 'Symbol' in df.columns:
# #             df.loc[mask, "Symbol"] = df.loc[mask, "Symbol"].astype(str).str[-5:] + df.loc[mask, "Symbol"].astype(str).str[-8:-6]
# #         else:
# #             raise KeyError("Symbol column not found in positions file")
# #         df_nfo = df[df["Exchange"] == "NFO"].copy()
# #         df_bfo = df[df["Exchange"] == "BFO"].copy()
# #         total_settlement_pnl_nfo = 0
# #         total_settlement_pnl_bfo = 0
# #         df_bhav_nfo = pd.DataFrame()
# #         df_bhav_bfo = pd.DataFrame()
# #         if include_settlement_nfo:
# #             if nfo_bhav_file is None:
# #                 raise ValueError("NFO bhavcopy file is required when settlement PNL for NFO is enabled")
# #             df_bhav_nfo = pd.read_csv(nfo_bhav_file)
# #             if 'CONTRACT_D' not in df_bhav_nfo.columns:
# #                 raise ValueError("CONTRACT_D column not found in NFO bhavcopy")
# #             df_bhav_nfo["Date"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(\d{2}-[A-Z]{3}-\d{4})')
# #             df_bhav_nfo["Symbol"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'^(.*?)(\d{2}-[A-Z]{3}-\d{4})')[0]
# #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(PE\d+|CE\d+)$')
# #             df_bhav_nfo["Date"] = pd.to_datetime(df_bhav_nfo["Date"], format="%d-%b-%Y")
# #             df_bhav_nfo["Strike_Type"] = df_bhav_nfo["Strike_Type"].str.replace(r'^(PE|CE)(\d+)$', r'\2\1', regex=True)
# #             target_symbol = "OPTIDXNIFTY"
# #             df_bhav_nfo = df_bhav_nfo[
# #                 (df_bhav_nfo["Date"] == pd.to_datetime(expiry_nfo)) &
# #                 (df_bhav_nfo["Symbol"] == target_symbol)
# #             ]
# #             df_nfo["Strike_Type"] = df_nfo["Symbol"].str.extract(r'(\d+[A-Z]{2})$')
# #             df_nfo['Strike'] = df_nfo['Strike_Type'].str[:-2].astype(float, errors='ignore')
# #             df_nfo['Option_Type'] = df_nfo['Strike_Type'].str[-2:]
# #             df_nfo = df_nfo.merge(
# #                 df_bhav_nfo[["Strike_Type", "SETTLEMENT"]],
# #                 on="Strike_Type",
# #                 how="left"
# #             )
# #         if include_settlement_bfo:
# #             if bfo_bhav_file is None:
# #                 raise ValueError("BFO bhavcopy file is required when settlement PNL for BFO is enabled")
# #             df_bhav_bfo = pd.read_csv(bfo_bhav_file)
# #             df_bhav_bfo["Date"] = pd.to_datetime(df_bhav_bfo["Market Summary Date"], format="%d %b %Y", errors="coerce")
# #             df_bhav_bfo["Expiry Date"] = pd.to_datetime(df_bhav_bfo["Expiry Date"], format="%d %b %Y", errors="coerce")
# #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Series Code"].astype(str).str[-7:]
# #             df_bhav_bfo = df_bhav_bfo[(df_bhav_bfo["Expiry Date"] == pd.to_datetime(expiry_bfo))]
# #             df_bfo["Symbol"] = df_bfo["Symbol"].astype(str).str.strip()
# #             df_bhav_bfo["Symbols"] = df_bhav_bfo["Symbols"].astype(str).str.strip()
# #             bhav_mapping = df_bhav_bfo.drop_duplicates(subset="Symbols", keep="last").set_index("Symbols")["Close Price"]
# #             df_bfo["Close Price"] = df_bfo["Symbol"].map(bhav_mapping)

# #         conditions = [
# #             df_nfo["Net Qty"] == 0,
# #             df_nfo["Net Qty"] > 0,
# #             df_nfo["Net Qty"] < 0
# #         ]
# #         choices = [
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
# #             (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Buy Qty"]
# #         ]
# #         df_nfo["Calculated_Realized_PNL"] = np.select(conditions, choices, default=0)
# #         df_nfo["Matching_Realized"] = df_nfo["Realized Profit"] == df_nfo["Calculated_Realized_PNL"]
# #         df_nfo["Matching_Realized"] = df_nfo["Matching_Realized"].replace({True: "TRUE", False: ""})
# #         if include_settlement_nfo:
# #             df_nfo["Calculated_Settlement_PNL"] = np.select(
# #                 [df_nfo["Net Qty"] > 0, df_nfo["Net Qty"] < 0],
# #                 [(df_nfo["SETTLEMENT"] - df_nfo["Buy Avg Price"]) * abs(df_nfo["Net Qty"]),
# #                  (df_nfo["Sell Avg Price"] - df_nfo["SETTLEMENT"]) * abs(df_nfo["Net Qty"])],
# #                 default=0
# #             )
# #             df_nfo["Matching_Settlement"] = df_nfo["Unrealized Profit"] == df_nfo["Calculated_Settlement_PNL"]
# #             df_nfo["Matching_Settlement"] = df_nfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# #             total_settlement_pnl_nfo = df_nfo["Calculated_Settlement_PNL"].fillna(0).sum()
# #         else:
# #             df_nfo["Calculated_Settlement_PNL"] = 0
# #             df_nfo["Matching_Settlement"] = ""
# #         total_realized_pnl_nfo = df_nfo["Calculated_Realized_PNL"].fillna(0).sum()

# #         conditions_bfo = [
# #             df_bfo["Net Qty"] == 0,
# #             df_bfo["Net Qty"] > 0,
# #             df_bfo["Net Qty"] < 0
# #         ]
# #         choices_bfo = [
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
# #             (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Buy Qty"]
# #         ]
# #         df_bfo["Calculated_Realized_PNL"] = np.select(conditions_bfo, choices_bfo, default=0)
# #         df_bfo["Matching_Realized"] = df_bfo["Realized Profit"] == df_bfo["Calculated_Realized_PNL"]
# #         df_bfo["Matching_Realized"] = df_bfo["Matching_Realized"].replace({True: "TRUE", False: ""})
# #         if include_settlement_bfo:
# #             long_condition = df_bfo["Net Qty"] > 0
# #             df_bfo.loc[long_condition, "Calculated_Settlement_PNL"] = (df_bfo["Close Price"] - df_bfo["Buy Avg Price"]) * abs(df_bfo["Net Qty"])
# #             short_condition = df_bfo["Net Qty"] < 0
# #             df_bfo.loc[short_condition, "Calculated_Settlement_PNL"] = (df_bfo["Sell Avg Price"] - df_bfo["Close Price"]) * abs(df_bfo["Net Qty"])
# #             df_bfo.loc[df_bfo["Net Qty"] == 0, "Calculated_Settlement_PNL"] = 0
# #             df_bfo["Matching_Settlement"] = df_bfo["Unrealized Profit"] == df_bfo["Calculated_Settlement_PNL"]
# #             df_bfo["Matching_Settlement"] = df_bfo["Matching_Settlement"].replace({True: "TRUE", False: ""})
# #             total_settlement_pnl_bfo = df_bfo["Calculated_Settlement_PNL"].fillna(0).sum()
# #         else:
# #             df_bfo["Calculated_Settlement_PNL"] = 0
# #             df_bfo["Matching_Settlement"] = ""
# #         total_realized_pnl_bfo = df_bfo["Calculated_Realized_PNL"].fillna(0).sum()

# #         overall_realized = total_realized_pnl_nfo + total_realized_pnl_bfo
# #         overall_settlement = total_settlement_pnl_nfo + total_settlement_pnl_bfo
# #         grand_total = overall_realized + overall_settlement

# #         return {
# #             "df_nfo": df_nfo,
# #             "df_bfo": df_bfo,
# #             "total_realized_nfo": total_realized_pnl_nfo,
# #             "total_settlement_nfo": total_settlement_pnl_nfo,
# #             "total_realized_bfo": total_realized_pnl_bfo,
# #             "total_settlement_bfo": total_settlement_pnl_bfo,
# #             "overall_realized": overall_realized,
# #             "overall_settlement": overall_settlement,
# #             "grand_total": grand_total
# #         }
# #     except Exception as e:
# #         logger.error(f"Error in process_data: {str(e)}")
# #         raise

# # # ===================== DOWNLOAD FUNCTIONS =====================
# # def get_excel_download_link(df, filename):
# #     output = BytesIO()
# #     with pd.ExcelWriter(output, engine='openpyxl') as writer:
# #         df.to_excel(writer, index=False, sheet_name='PNL Data')
# #         worksheet = writer.sheets['PNL Data']
# #         for row in worksheet.rows:
# #             for cell in row:
# #                 cell.alignment = Alignment(horizontal='center')
# #         for cell in worksheet[1]:
# #             cell.font = Font(bold=True, color="FFFFFF")
# #             cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
# #         for row in worksheet.rows:
# #             for cell in row:
# #                 cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
# #     excel_data = output.getvalue()
# #     b64 = base64.b64encode(excel_data).decode()
# #     return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}.xlsx</a>'

# # def get_csv_download_link(df, filename):
# #     output = BytesIO()
# #     df.to_csv(output, index=False)
# #     csv_data = output.getvalue()
# #     b64 = base64.b64encode(csv_data).decode()
# #     return f'<a href="data:text/csv;base64,{b64}" download="{filename}" class="text-blue-600 hover:text-blue-800 font-semibold">Download {filename}</a>'

# # ===================== RUN THE APP =====================
# if __name__ == "__main__":
#     st.write(f"DEBUG: Starting Streamlit app at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")





import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
from io import BytesIO
import logging
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Check Streamlit version for compatibility
try:
    import streamlit
    logger.info(f"Streamlit version: {streamlit.__version__}")
except ImportError:
    st.error("Streamlit is not installed. Please install it using `pip install streamlit`.")
    st.stop()


# ===================== MAIN RUN FUNCTION =====================
def run():
    # ===================== CUSTOM CSS & STYLING =====================
    st.markdown("""
    <link href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css" rel="stylesheet">
    <style>
    body {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #f3f4f6, #e5e7eb);
    }
    .stButton>button {
        background: linear-gradient(45deg, #3b82f6, #60a5fa);
        color: white;
        border: none;
        padding: 0.75rem 1.5rem;
        border-radius: 0.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
        width: 100%;
    }
    .stButton>button:hover {
        background: linear-gradient(45deg, #2563eb, #3b82f6);
        transform: translateY(-2px);
        box-shadow: 0 4px 6px rgba(0,0,0,0.2);
    }
    .stDateInput input {
        border: 2px solid #3b82f6;
        border-radius: 0.5rem;
        padding: 0.5rem;
        background: #ffffff;
        color: #1f2937;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    .stDateInput input:focus {
        outline: none;
        border-color: #2563eb;
        box-shadow: 0 0 8px rgba(59, 130, 246, 0.5);
        background: #f8fafc;
    }
    .stFileUploader button {
        background: linear-gradient(45deg, #10b981, #34d399);
        color: white;
        border-radius: 0.5rem;
        padding: 0.75rem;
    }
    .stFileUploader button:hover {
        background: linear-gradient(45deg, #059669, #10b981);
    }
    .stCheckbox label {
        font-size: 1rem;
        color: #1f2937;
    }
    .metric-card {
        background: #ffffff;
        padding: 1.5rem;
        border-radius: 0.75rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        text-align: center;
        transition: transform 0.3s ease;
    }
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 6px 12px rgba(0,0,0,0.15);
    }
    .metric-label {
        font-size: 1.1rem;
        color: #6b7280;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 1.75rem;
        font-weight: 700;
    }
    .stTabs [data-baseweb="tab"] {
        font-size: 1.1rem;
        font-weight: 600;
        padding: 0.75rem 1.5rem;
        border-radius: 0.5rem;
        transition: all 0.3s ease;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background: #e5e7eb;
    }
    .insights-box {
        background: #ffffff;
        padding: 1.5rem;
        border-radius: 0.5rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-top: 1.5rem;
    }
    .chart-container {
        background: #ffffff;
        padding: 1.5rem;
        border-radius: 0.75rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-bottom: 1.5rem;
    }
    .header-text {
        font-size: 2.25rem;
        font-weight: 800;
        color: #1f2937;
        text-align: center;
        margin-bottom: 1rem;
    }
    .subheader-text {
        font-size: 1.25rem;
        color: #4b5563;
        text-align: center;
        margin-bottom: 2rem;
    }
    footer { visibility: hidden; }
    </style>
    """, unsafe_allow_html=True)

    # ===================== ALL FUNCTIONS FIRST (EXACT SAME) =====================

    def process_portfolio_data(gridlog_file, summary_file):
        if gridlog_file.name.endswith('.csv'):
            df_grid = pd.read_csv(gridlog_file)
        elif gridlog_file.name.endswith('.xlsx'):
            df_grid = pd.read_excel(gridlog_file)
        else:
            raise ValueError("Unsupported GridLog file type. Use CSV or Excel.")

        df_grid.columns = df_grid.columns.str.strip()

        mask = df_grid['Message'].str.contains(r'Combined SL:|Combined trail target:', case=False, na=False)
        filtered_grid = df_grid.loc[mask, ['Message', 'Option Portfolio', 'Timestamp']].dropna(subset=['Option Portfolio'])

        filtered_grid['MessageType'] = filtered_grid['Message'].str.extract(r'(Combined SL|Combined trail target)', flags=re.IGNORECASE)
        duplicate_mask = filtered_grid.duplicated(subset=['Option Portfolio', 'MessageType'], keep=False)
        filtered_grid = filtered_grid[duplicate_mask]

        summary_grid = (
            filtered_grid.groupby('Option Portfolio').agg({
                'Message': lambda x: ', '.join(x.unique()),
                'Timestamp': 'max'
            }).reset_index()
            .rename(columns={'Message': 'Reason', 'Timestamp': 'Time'})
        )

        xl = pd.ExcelFile(summary_file)
        summary_list = []

        for sheet_name in xl.sheet_names:
            if "legs" in sheet_name.lower():
                df_leg = xl.parse(sheet_name)
                df_leg.columns = df_leg.columns.str.strip()

                if {'Exit Type', 'Portfolio Name', 'Exit Time'}.issubset(df_leg.columns):
                    onsqoff_df = df_leg[df_leg['Exit Type'].astype(str).str.strip() == 'OnSqOffTime']
                    if not onsqoff_df.empty:
                        grouped = onsqoff_df.groupby('Portfolio Name')['Exit Time'].max().reset_index()
                        for _, row in grouped.iterrows():
                            summary_list.append({
                                'Option Portfolio': row['Portfolio Name'],
                                'Reason': 'OnSqOffTime',
                                'Time': row['Exit Time']
                            })

        summary_summary = pd.DataFrame(summary_list)
        final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)
        final_df = final_df.groupby('Option Portfolio').agg({
            'Reason': lambda x: ', '.join(sorted(set(x))),
            'Time': 'last'
        }).reset_index()

        completed_list = []
        grid_portfolios = df_grid['Option Portfolio'].dropna().unique()

        for sheet_name in xl.sheet_names:
            if "legs" in sheet_name.lower():
                df_leg = xl.parse(sheet_name)
                df_leg.columns = df_leg.columns.str.strip()

                if 'Portfolio Name' in df_leg.columns and 'Status' in df_leg.columns:
                    for portfolio, group in df_leg.groupby('Portfolio Name'):
                        if (portfolio not in final_df['Option Portfolio'].values 
                            and portfolio in grid_portfolios):
                            statuses = group['Status'].astype(str).str.strip().unique()
                            if len(statuses) == 1 and statuses[0].lower() == 'completed':
                                reason_text = 'AllLegsCompleted'
                                exit_time_to_use = None
                                if 'Exit Time' in group.columns:
                                    for exit_time, exit_type in zip(group['Exit Time'], group.get('Exit Type', [])):
                                        if pd.isna(exit_time):
                                            continue
                                        normalized_exit_time = str(exit_time).replace('.', ':').strip()
                                        matching_rows = df_grid[
                                            (df_grid['Option Portfolio'] == portfolio) &
                                            (df_grid['Timestamp'].astype(str).str.contains(normalized_exit_time))
                                        ]
                                        if not matching_rows.empty:
                                            reason_text += f", {exit_type.strip()}"
                                            exit_time_to_use = exit_time
                                            break
                                completed_list.append({
                                    'Option Portfolio': portfolio,
                                    'Reason': reason_text,
                                    'Time': exit_time_to_use
                                })

        if completed_list:
            completed_df = pd.DataFrame(completed_list)
            final_df = pd.concat([final_df, completed_df], ignore_index=True)

        def clean_reason(text):
            if pd.isna(text):
                return text
            text = str(text)
            match = re.search(r'(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)', text, re.IGNORECASE)
            if match:
                return match.group(1)
            if 'AllLegsCompleted' in text:
                text = text.replace('AllLegsCompleted,', '').replace('AllLegsCompleted', '').strip()
            return text.strip()

        final_df['Reason'] = final_df['Reason'].apply(clean_reason)

        filename = gridlog_file.name
        match = re.search(r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', filename)
        if match:
            raw_date = match.group(1)
            parts = raw_date.split()
            formatted_date = f"{parts[0]} {parts[1].lower()}"
        else:
            formatted_date = "unknown_date"
        output_filename = f"completed portfolio of {formatted_date}.csv"

        final_df['Time'] = final_df['Time'].astype(str).str.strip().replace('nan', None)
        return final_df, output_filename

    def process_data(df, nfo_bhav_file, bfo_bhav_file, expiry_nfo, expiry_bfo,
                     include_settlement_nfo, include_settlement_bfo):
        logger.info("Starting PNL data processing")
        try:
            required_columns = ['Exchange', 'Symbol', 'Net Qty', 'Buy Avg Price', 'Sell Avg Price',
                                'Sell Qty', 'Buy Qty', 'Realized Profit', 'Unrealized Profit']
            missing = [c for c in required_columns if c not in df.columns]
            if missing:
                raise ValueError(f"Missing columns: {missing}")

            mask = df["Exchange"].isin(["NFO", "BFO"])
            df.loc[mask, "Symbol"] = df.loc[mask, "Symbol"].astype(str).str[-5:] + df.loc[mask, "Symbol"].astype(str).str[-8:-6]

            df_nfo = df[df["Exchange"] == "NFO"].copy()
            df_bfo = df[df["Exchange"] == "BFO"].copy()

            total_realized_nfo = total_realized_bfo = 0
            total_settlement_nfo = total_settlement_bfo = 0

            cond_nfo = [df_nfo["Net Qty"] == 0, df_nfo["Net Qty"] > 0, df_nfo["Net Qty"] < 0]
            choice_nfo = [
                (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
                (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Sell Qty"],
                (df_nfo["Sell Avg Price"] - df_nfo["Buy Avg Price"]) * df_nfo["Buy Qty"]
            ]
            df_nfo["Calculated_Realized_PNL"] = np.select(cond_nfo, choice_nfo, default=0)
            total_realized_nfo = df_nfo["Calculated_Realized_PNL"].fillna(0).sum()

            cond_bfo = [df_bfo["Net Qty"] == 0, df_bfo["Net Qty"] > 0, df_bfo["Net Qty"] < 0]
            choice_bfo = [
                (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
                (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Sell Qty"],
                (df_bfo["Sell Avg Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Buy Qty"]
            ]
            df_bfo["Calculated_Realized_PNL"] = np.select(cond_bfo, choice_bfo, default=0)
            total_realized_bfo = df_bfo["Calculated_Realized_PNL"].fillna(0).sum()

            if include_settlement_nfo and nfo_bhav_file:
                df_bhav_nfo = pd.read_csv(nfo_bhav_file)
                df_bhav_nfo["Date"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(\d{2}-[A-Z]{3}-\d{4})')
                df_bhav_nfo["Symbol"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'^(.*?)(\d{2}-[A-Z]{3}-\d{4})')[0]
                df_bhav_nfo["Strike_Type"] = df_bhav_nfo["CONTRACT_D"].str.extract(r'(PE\d+|CE\d+)$')
                df_bhav_nfo["Date"] = pd.to_datetime(df_bhav_nfo["Date"], format="%d-%b-%Y")
                df_bhav_nfo["Strike_Type"] = df_bhav_nfo["Strike_Type"].str.replace(r'^(PE|CE)(\d+)$', r'\2\1', regex=True)
                df_bhav_nfo = df_bhav_nfo[(df_bhav_nfo["Date"] == pd.to_datetime(expiry_nfo)) & (df_bhav_nfo["Symbol"] == "OPTIDXNIFTY")]
                df_nfo["Strike_Type"] = df_nfo["Symbol"].str.extract(r'(\d+[A-Z]{2})$')
                df_nfo = df_nfo.merge(df_bhav_nfo[["Strike_Type", "SETTLEMENT"]], on="Strike_Type", how="left")
                df_nfo["Calculated_Settlement_PNL"] = np.select(
                    [df_nfo["Net Qty"] > 0, df_nfo["Net Qty"] < 0],
                    [(df_nfo["SETTLEMENT"] - df_nfo["Buy Avg Price"]) * abs(df_nfo["Net Qty"]),
                     (df_nfo["Sell Avg Price"] - df_nfo["SETTLEMENT"]) * abs(df_nfo["Net Qty"])],
                    default=0)
                total_settlement_nfo = df_nfo["Calculated_Settlement_PNL"].fillna(0).sum()

            if include_settlement_bfo and bfo_bhav_file:
                df_bhav_bfo = pd.read_csv(bfo_bhav_file)
                df_bhav_bfo["Expiry Date"] = pd.to_datetime(df_bhav_bfo["Expiry Date"], format="%d %b %Y", errors="coerce")
                df_bhav_bfo = df_bhav_bfo[df_bhav_bfo["Expiry Date"] == pd.to_datetime(expiry_bfo)]
                df_bhav_bfo["Symbols"] = df_bhav_bfo["Series Code"].astype(str).str[-7:]
                mapping = df_bhav_bfo.drop_duplicates("Symbols").set_index("Symbols")["Close Price"]
                df_bfo["Close Price"] = df_bfo["Symbol"].astype(str).str.strip().map(mapping)
                df_bfo["Calculated_Settlement_PNL"] = 0
                df_bfo.loc[df_bfo["Net Qty"] > 0, "Calculated_Settlement_PNL"] = (df_bfo["Close Price"] - df_bfo["Buy Avg Price"]) * df_bfo["Net Qty"].abs()
                df_bfo.loc[df_bfo["Net Qty"] < 0, "Calculated_Settlement_PNL"] = (df_bfo["Sell Avg Price"] - df_bfo["Close Price"]) * df_bfo["Net Qty"].abs()
                total_settlement_bfo = df_bfo["Calculated_Settlement_PNL"].fillna(0).sum()

            overall_realized = total_realized_nfo + total_realized_bfo
            overall_settlement = total_settlement_nfo + total_settlement_bfo
            grand_total = overall_realized + overall_settlement

            return {
                "total_realized_nfo": total_realized_nfo,
                "total_settlement_nfo": total_settlement_nfo,
                "total_realized_bfo": total_realized_bfo,
                "total_settlement_bfo": total_settlement_bfo,
                "overall_realized": overall_realized,
                "overall_settlement": overall_settlement,
                "grand_total": grand_total
            }
        except Exception as e:
            logger.error(f"Error in process_data: {e}")
            raise

    def get_excel_download_link(df, filename):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PNL Data')
            ws = writer.sheets['PNL Data']
            for row in ws.rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", fill_type="solid")
        b64 = base64.b64encode(output.getvalue()).decode()
        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Download {filename}.xlsx</a>'

    def get_csv_download_link(df, filename):
        csv = df.to_csv(index=False).encode()
        b64 = base64.b64encode(csv).decode()
        return f'<a href="data:text/csv;base64,{b64}" download="{filename}">Download {filename}</a>'

    # ===================== TABS =====================
    tab1, tab2 = st.tabs(["Full PNL Calculation", "Portfolio Exit Analysis"])

   # ===================== TAB 1: PNL CALCULATION =====================
    with tab1:
        st.markdown('<h1 class="header-text">Ultimate Financial PNL Dashboard</h1>', unsafe_allow_html=True)
        st.markdown('<p class="subheader-text">Analyze your portfolio with real-time insights, interactive charts, and comprehensive data exports for smarter trading decisions.</p>', unsafe_allow_html=True)
       
        st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Upload Data</h2>', unsafe_allow_html=True)
        with st.container():
            positions_file = st.file_uploader("Positions CSV", type="csv", help="Upload VS20 22 AUG 2025 POSITIONS(EOD).csv", key=("positions_upload"))
            selected_user = None
            if positions_file:
                if 'positions_df' not in st.session_state or st.session_state.get('positions_file_name') != positions_file.name:
                    st.session_state.positions_df = pd.read_csv(positions_file)
                    st.session_state.positions_file_name = positions_file.name
                df = st.session_state.positions_df
                if 'UserID' in df.columns:
                    users = sorted(df['UserID'].unique().tolist())
                    selected_user = st.selectbox("Select User", users, key="selected_user")
                else:
                    st.error("'UserID' column not found in positions file.")
           
            checkbox_col1, checkbox_col2 = st.columns(2)
            with checkbox_col1:
                include_settlement_nfo = st.checkbox("Include Settlement PNL for NFO", value=True, key="nfo_settlement")
            with checkbox_col2:
                include_settlement_bfo = st.checkbox("Include Settlement PNL for BFO", value=True, key="bfo_settlement")
           
            col1, col2 = st.columns(2)
            with col1:
                nfo_bhav_file = st.file_uploader("NFO Bhavcopy", type="csv", key="nfo_upload") if include_settlement_nfo else None
                if not include_settlement_nfo:
                    st.info("NFO Bhavcopy not required when settlement PNL for NFO is disabled.")
            with col2:
                bfo_bhav_file = st.file_uploader("BFO Bhavcopy", type="csv", key="bfo_upload") if include_settlement_bfo else None
                if not include_settlement_bfo:
                    st.info("BFO Bhavcopy not required when settlement PNL for BFO is disabled.")
           
            st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Expiry Dates</h2>', unsafe_allow_html=True)
            col3, col4 = st.columns(2)
            with col3:
                expiry_nfo = st.date_input("NFO Expiry Date", value=datetime.now().date(), key="nfo_expiry", disabled=not include_settlement_nfo)
            with col4:
                expiry_bfo = st.date_input("BFO Expiry Date", value=datetime.now().date(), key="bfo_expiry", disabled=not include_settlement_bfo)
           
            process_button = st.button("Process Data", key="process_button")

            if process_button:
                if positions_file and selected_user:
                    if (include_settlement_nfo and not nfo_bhav_file) or (include_settlement_bfo and not bfo_bhav_file):
                        st.error("Please upload all required files.")
                    else:
                        try:
                            with st.spinner("Processing PNL..."):
                                filtered_df = st.session_state.positions_df[st.session_state.positions_df['UserID'] == selected_user]
                                results = process_data(
                                    filtered_df, nfo_bhav_file, bfo_bhav_file,
                                    expiry_nfo, expiry_bfo,
                                    include_settlement_nfo, include_settlement_bfo
                                )

                            st.success("PNL processed successfully!")

                            # ==================== DISPLAY RESULTS WITH METRIC CARDS ====================
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">NFO Realized PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['total_realized_nfo'] >= 0 else '#ef4444'}">
                                        ₹{results['total_realized_nfo']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            with col2:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">NFO Settlement PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['total_settlement_nfo'] >= 0 else '#ef4444'}">
                                        ₹{results['total_settlement_nfo']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            with col3:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">BFO Realized PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['total_realized_bfo'] >= 0 else '#ef4444'}">
                                        ₹{results['total_realized_bfo']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            with col4:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">BFO Settlement PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['total_settlement_bfo'] >= 0 else '#ef4444'}">
                                        ₹{results['total_settlement_bfo']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)

                            # Overall Summary
                            st.markdown("### Overall Summary")
                            colA, colB, colC = st.columns(3)
                            with colA:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">Total Realized PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['overall_realized'] >= 0 else '#ef4444'}; font-size: 2rem;">
                                        ₹{results['overall_realized']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            with colB:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">Total Settlement PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['overall_settlement'] >= 0 else '#ef4444'}; font-size: 2rem;">
                                        ₹{results['overall_settlement']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
                            with colC:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">Grand Total PNL</div>
                                    <div class="metric-value" style="color: {'#10b981' if results['grand_total'] >= 0 else '#ef4444'}; font-size: 2.5rem;">
                                        ₹{results['grand_total']:,.2f}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)

                            # ==================== DOWNLOAD FILTERED DATA ====================
                            st.markdown("### Download Processed Positions Data")
                            download_df = filtered_df.copy()
                            csv_link = get_csv_download_link(download_df, f"PNL_{selected_user}_{datetime.now().strftime('%Y%m%d')}.csv")
                            excel_link = get_excel_download_link(download_df, f"PNL_{selected_user}_{datetime.now().strftime('%Y%m%d')}")

                            col_d1, col_d2 = st.columns(2)
                            with col_d1:
                                st.markdown(csv_link, unsafe_allow_html=True)
                            with col_d2:
                                st.markdown(excel_link, unsafe_allow_html=True)

                        except Exception as e:
                            st.error(f"Error during processing: {e}")
                            logger.error(f"Processing error: {e}", exc_info=True)
                else:
                    st.error("Please upload positions file and select a user.")

    # ===================== TAB 2: PORTFOLIO ANALYSIS =====================
    with tab2:
        st.markdown('<hr class="my-8 border-gray-300">', unsafe_allow_html=True)
        st.markdown('<h1 class="header-text">Portfolio Analysis</h1>', unsafe_allow_html=True)
        st.markdown('<p class="subheader-text">Upload GridLog and Summary files to analyze portfolio exit reasons and timestamps.</p>', unsafe_allow_html=True)
        st.info("Upload the required files below and click 'Process Portfolio Data' to view results.")
        st.markdown('<h2 class="text-lg font-bold text-gray-800 dark:text-gray-200 mb-4">Upload Portfolio Data</h2>', unsafe_allow_html=True)
        
        col_grid, col_summary = st.columns(2)
        with col_grid:
            gridlog_file = st.file_uploader("GridLog File", type=["csv", "xlsx"], key="gridlog_upload")
        with col_summary:
            summary_file = st.file_uploader("Summary Excel File", type="xlsx", key="summary_upload")
        
        if st.button("Process Portfolio Data", key="process_portfolio_button"):
            if gridlog_file and summary_file:
                try:
                    with st.spinner("Processing portfolio data..."):
                        final_df, output_filename = process_portfolio_data(gridlog_file, summary_file)
                    st.success("Done!")
                    st.write(final_df)
                    st.markdown(get_csv_download_link(final_df, output_filename), unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Error: {e}")
            else:
                st.error("Please upload both files.")

# ===================== AUTO CALL run() =====================
if __name__ == "__main__":
    st.write(f"DEBUG: Starting app at {datetime.now()}")
    run()

